import os
import json
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
from io import BytesIO, StringIO
from datetime import datetime
from zoneinfo import ZoneInfo

_VAN = ZoneInfo("America/Vancouver")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

PRIMARY_LOOKUP_PATH = "institution_lookup_primary.xlsx"
LOOKUP_PATH         = "institution_lookup.xlsx"   # backup
HISTORY_PATH       = os.path.join(os.path.expanduser("~"), ".ratesheet", "special_rates_history.json")
TABLE_STYLE_PATH   = os.path.join(os.path.expanduser("~"), ".ratesheet", "table_style.json")
APP_SETTINGS_PATH  = os.path.join(os.path.expanduser("~"), ".ratesheet", "app_settings.json")

_DEFAULT_EMAIL_TEMPLATE = """Hi All,

[NOTES]

High-Interest Savings Account (CAD)
[HISA_CAD]

High-Interest Savings Account (USD)
[HISA_USD]

Guaranteed Investment Certificates (CAD)
[GIC_CAD_TABLE]

Guaranteed Investment Certificates (USD)
[GIC_USD_TABLE]

Thanks,"""

_DEFAULT_APP_SETTINGS = {
    "show_master_data":   True,
    "show_custom_query":  True,
    "show_rate_sheet":    True,
    "announcement":       "",
    "email_template":     _DEFAULT_EMAIL_TEMPLATE,
    "email_font":         "Calibri",
    "email_font_size":    11,
    "email_text_color":   "#000000",
}

@st.cache_resource
def _app_settings_store():
    store = dict(_DEFAULT_APP_SETTINGS)
    try:
        if os.path.exists(APP_SETTINGS_PATH):
            with open(APP_SETTINGS_PATH) as f:
                store.update(json.load(f))
    except Exception:
        pass
    return store

def _save_app_settings():
    store = _app_settings_store()
    try:
        os.makedirs(os.path.dirname(APP_SETTINGS_PATH), exist_ok=True)
        with open(APP_SETTINGS_PATH, "w") as f:
            json.dump(dict(store), f, indent=2)
    except Exception:
        pass

DEFAULT_TABLE_STYLE = {
    "header_bg":      "#000000",
    "header_text":    "#ffffff",
    "header_font":    "Calibri",
    "header_size":    11,
    "header_bold":    True,
    "body_font":      "Calibri",
    "body_size":      11,
    "body_text":      "#000000",
    "body_bg":        "#ffffff",
    "alt_row_bg":     "#ffffff",
    "rate_color":     "#C00000",
    "border_color":   "#cccccc",
    "border_width":   1,
    "cell_padding":   6,
}

FONT_OPTIONS = ["Calibri", "Arial", "Helvetica", "Times New Roman",
                "Georgia", "Verdana", "Trebuchet MS", "Tahoma"]

@st.cache_resource
def _table_style_store():
    store = dict(DEFAULT_TABLE_STYLE)
    try:
        if os.path.exists(TABLE_STYLE_PATH):
            with open(TABLE_STYLE_PATH) as f:
                store.update(json.load(f))
    except Exception:
        pass
    return store

def load_table_style():
    return dict(_table_style_store())

def save_table_style(updates):
    store = _table_style_store()
    store.update(updates)
    try:
        os.makedirs(os.path.dirname(TABLE_STYLE_PATH), exist_ok=True)
        with open(TABLE_STYLE_PATH, "w") as f:
            json.dump(dict(store), f, indent=2)
    except Exception:
        pass

def _style_preview_html(s):
    bw   = s["border_width"]
    pad  = s["cell_padding"]
    border      = f"{bw}px solid {s['border_color']}"
    # Header borders match the header background so lines are invisible
    hdr_border  = f"{bw}px solid {s['header_bg']}"
    th_style = (
        f"background:{s['header_bg']};color:{s['header_text']};"
        f"font-family:{s['header_font']},sans-serif;font-size:{s['header_size']}pt;"
        f"font-weight:{'bold' if s['header_bold'] else 'normal'};"
        f"border:{hdr_border};padding:{pad}px {pad*2}px;text-align:center;"
    )
    def td_style(i, is_rate=False):
        bg = s["alt_row_bg"] if i % 2 == 1 else s["body_bg"]
        color = s["rate_color"] if is_rate else s["body_text"]
        return (
            f"background:{bg};color:{color};"
            f"font-family:{s['body_font']},sans-serif;font-size:{s['body_size']}pt;"
            f"border:{border};padding:{pad}px {pad*2}px;text-align:center;"
        )
    rows = [
        ("Royal Bank of Canada",  "R-1 (High) – CDIC", "1 Year Fixed",  "3.75%"),
        ("TD Bank",               "R-1 (High) – CDIC", "1 Year Fixed",  "3.70%"),
        ("Oaken Financial",       "100% Guarantee – CDIC", "2 Year Fixed", "4.10%"),
        ("EQ Bank",               "BBB (High) – CDIC", "2 Year Fixed",  "4.05%"),
    ]
    body = ""
    for i, (issuer, rating, term, rate) in enumerate(rows):
        body += (
            f"<tr>"
            f"<td style='{td_style(i)}'>{issuer}</td>"
            f"<td style='{td_style(i)}'>{rating}</td>"
            f"<td style='{td_style(i)}'>{term}</td>"
            f"<td style='{td_style(i, is_rate=True)}'>{rate}</td>"
            f"</tr>"
        )
    return (
        f"<table style='border-collapse:collapse;width:100%;'>"
        f"<thead><tr>"
        f"<th style='{th_style}'>Issuer</th>"
        f"<th style='{th_style}'>Credit Rating & Guarantee</th>"
        f"<th style='{th_style}'>Term</th>"
        f"<th style='{th_style}'>Rate</th>"
        f"</tr></thead>"
        f"<tbody>{body}</tbody>"
        f"</table>"
    )


def format_usd_term(term):
    """Format USD term names: 30 → 30 days fixed, 1 → 1 year fixed, etc."""
    term = str(term).strip()
    if term == "30":
        return "30 days fixed"
    elif term == "60":
        return "60 days fixed"
    elif term == "90":
        return "90 days fixed"
    elif term == "120":
        return "120 days fixed"
    elif term == "180":
        return "180 days fixed"
    elif term == "270":
        return "270 days fixed"
    elif term == "1":
        return "1 year fixed"
    elif term == "18 months":
        return "18 months fixed"
    elif term == "2":
        return "2 years fixed"
    elif term.startswith("Cashable after"):
        # "Cashable after 30" → "Cashable after 30 days"
        return term + " days" if "days" not in term else term
    return term


def build_copy_html(rows, style=None):
    """Outlook-compatible HTML table with exact sizing and reliable colours."""
    s      = style or load_table_style()
    bw     = s["border_width"]
    pad    = s["cell_padding"]
    h_bg   = s["header_bg"]
    h_txt  = s["header_text"]
    h_fnt  = s["header_font"]
    h_sz   = s["header_size"]
    h_bold = s["header_bold"]
    b_fnt  = s["body_font"]
    b_sz   = s["body_size"]
    b_txt  = s["body_text"]
    b_bg   = s["body_bg"]
    alt_bg = s["alt_row_bg"]
    r_col  = s["rate_color"]
    bdr_c  = s["border_color"]

    # ── Dimensions (Outlook uses pt for height, px for width) ──────────────
    # 0.56 cm = 15.87 pt = 21 px
    ROW_H_PT  = "15.87pt"
    ROW_H_PX  = "21"
    # Column widths: 1 cm ≈ 37.795 px
    W_CR   = "283"   # Credit Rating & Guarantee: 7.49 cm
    W_TERM = "208"   # Term: 5.5 cm
    W_RATE = "132"   # Rate: 3.5 cm
    # Issuer: no fixed width — nowrap so it fits content on one line

    bdr_body = f"{bw}px solid {bdr_c}"
    bdr_hdr  = f"{bw}px solid {h_bg}"   # invisible header borders

    def row_bg(ri):
        return alt_bg if ri % 2 == 1 else b_bg

    # ── Colour helper ───────────────────────────────────────────────────────
    # mso-color-alt:windowtext tells Word's renderer to use the explicit hex
    # colour rather than substituting a theme colour (the root cause of rates
    # showing as black instead of red).
    def _span(text, color):
        return (
            f"<span style='color:{color};mso-color-alt:windowtext;'>"
            f"<font face='{b_fnt}' color='{color}'>{text}</font>"
            f"</span>"
        )

    # ── Cell content wrapper ────────────────────────────────────────────────
    # margin:0cm + margin-bottom:.0001pt + mso-pagination:none is the exact
    # format Word uses internally for compact table cells. Without these,
    # Outlook adds paragraph spacing that pushes rows beyond the set height.
    _P_BASE = (
        "margin:0cm;margin-bottom:.0001pt;padding:0;"
        "mso-pagination:none;mso-line-height-rule:exactly;"
        "text-align:center;"
    )

    def _p(content, color):
        return (
            f"<p align='center' style='{_P_BASE}color:{color};'>"
            + _span(content, color) +
            "</p>"
        )

    def _p_hdr(content):
        b_o = "<b>" if h_bold else ""
        b_c = "</b>" if h_bold else ""
        return (
            f"<p align='center' style='{_P_BASE}color:{h_txt};'>"
            f"<span style='color:{h_txt};mso-color-alt:windowtext;'>"
            f"<font face='{h_fnt}' color='{h_txt}'><u>{b_o}{content}{b_c}</u></font>"
            f"</span></p>"
        )

    # ── Linkify insurance provider name ─────────────────────────────────────
    def linkify_outlook(text):
        if not text or text == "* CANNOT SOURCE, ENTER MANUALLY *":
            return text or ""
        if " – " in text:
            rating_part, ins_part = text.split(" – ", 1)
            provider, url = find_insurance_match(ins_part)
            if url:
                return f"{rating_part} – {_make_link(provider, url, ins_part)}"
            return text
        provider, url = find_insurance_match(text)
        return _make_link(provider, url, text) if url else text

    # ── Shared cell CSS ─────────────────────────────────────────────────────
    def th_css(width=""):
        w = f"width:{width}px;" if width else ""
        return (
            f"background-color:{h_bg};color:{h_txt};"
            f"font-family:{h_fnt},sans-serif;font-size:{h_sz}pt;"
            f"font-weight:{'bold' if h_bold else 'normal'};"
            f"border:{bdr_hdr};padding:2px {pad*2}px;"
            f"text-align:center;vertical-align:middle;"
            f"height:{ROW_H_PT};mso-line-height-rule:exactly;{w}"
        )

    def td_css(ri, color, width="", nowrap=False):
        bg = row_bg(ri)
        w  = f"width:{width}px;" if width else ""
        nw = "white-space:nowrap;" if nowrap else ""
        return (
            f"background-color:{bg};color:{color};"
            f"font-family:{b_fnt},sans-serif;font-size:{b_sz}pt;"
            f"border:{bdr_body};padding:2px {pad*2}px;"
            f"text-align:center;vertical-align:middle;"
            f"height:{ROW_H_PT};mso-line-height-rule:exactly;{w}{nw}"
        )

    # ── Term rowspans ───────────────────────────────────────────────────────
    rowspans, i = [], 0
    while i < len(rows):
        span = 1
        while i + span < len(rows) and rows[i + span][2] == rows[i][2]:
            span += 1
        rowspans.extend([span] + [0] * (span - 1))
        i += span

    # ── Build HTML ──────────────────────────────────────────────────────────
    def th_cell(label, width=""):
        w_attr = f" width='{width}'" if width else ""
        return (
            f"<th{w_attr} bgcolor='{h_bg}' align='center' height='{ROW_H_PX}' "
            f"style='{th_css(width)}'>"
            + _p_hdr(label) + "</th>"
        )

    html = (
        "<table border='0' cellpadding='0' cellspacing='0' "
        "style='border-collapse:collapse;'>"
        "<thead>"
        f"<tr height='{ROW_H_PX}'>"
        + th_cell("Issuer")
        + th_cell("Credit Rating &amp; Guarantee", W_CR)
        + th_cell("Term", W_TERM)
        + th_cell("Rate", W_RATE)
        + "</tr></thead><tbody>"
    )

    for ri, (issuer, rating, term, rate) in enumerate(rows):
        span     = rowspans[ri]
        rate_str = f"{rate * 100:.2f}%"
        bg       = row_bg(ri)

        html += f"<tr height='{ROW_H_PX}'>"

        # Issuer — no fixed width, nowrap
        html += (
            f"<td bgcolor='{bg}' align='center' valign='middle' "
            f"style='{td_css(ri, b_txt, nowrap=True)}'>"
            + _p(issuer, b_txt) + "</td>"
        )
        # Credit Rating
        html += (
            f"<td width='{W_CR}' bgcolor='{bg}' align='center' valign='middle' "
            f"style='{td_css(ri, b_txt, W_CR)}'>"
            + _p(linkify_outlook(str(rating)), b_txt) + "</td>"
        )
        # Term (rowspan)
        if span > 0:
            rs = f" rowspan='{span}'" if span > 1 else ""
            html += (
                f"<td{rs} width='{W_TERM}' bgcolor='{bg}' align='center' valign='middle' "
                f"style='{td_css(ri, b_txt, W_TERM)}'>"
                + _p(term, b_txt) + "</td>"
            )
        # Rate
        html += (
            f"<td width='{W_RATE}' bgcolor='{bg}' align='center' valign='middle' "
            f"style='{td_css(ri, r_col, W_RATE)}'>"
            + _p(rate_str, r_col) + "</td>"
        )
        html += "</tr>"

    html += "</tbody></table>"
    return html


def _build_simple_table(headers, rows, style=None):
    """Outlook-compatible table with fixed columns and no rowspans (for placeholder sections)."""
    s        = style or load_table_style()
    bw       = s["border_width"]
    h_bg     = s["header_bg"];  h_txt = s["header_text"]
    h_fnt    = s["header_font"]; h_sz = s["header_size"]; h_bold = s["header_bold"]
    b_fnt    = s["body_font"];  b_sz  = s["body_size"];  b_txt  = s["body_text"]
    b_bg     = s["body_bg"];    alt   = s["alt_row_bg"]; bdr_c  = s["border_color"]
    H_PT     = "15.87pt";       H_PX  = "21"
    hdr_bdr  = f"{bw}px solid {h_bg}"
    bdy_bdr  = f"{bw}px solid {bdr_c}"
    _PB      = ("margin:0cm;margin-bottom:.0001pt;padding:0;mso-pagination:none;"
                "mso-line-height-rule:exactly;text-align:center;")
    b_o = "<b>" if h_bold else ""; b_c = "</b>" if h_bold else ""

    th_css = (f"background-color:{h_bg};color:{h_txt};font-family:{h_fnt},sans-serif;"
              f"font-size:{h_sz}pt;font-weight:{'bold' if h_bold else 'normal'};"
              f"border:{hdr_bdr};padding:2px 12px;text-align:center;"
              f"height:{H_PT};mso-line-height-rule:exactly;")

    def td_css(ri):
        bg = alt if ri % 2 == 1 else b_bg
        return (f"background-color:{bg};color:{b_txt};font-family:{b_fnt},sans-serif;"
                f"font-size:{b_sz}pt;border:{bdy_bdr};padding:2px 12px;"
                f"text-align:center;vertical-align:middle;height:{H_PT};"
                f"mso-line-height-rule:exactly;")

    def cell_p(text, color, face):
        return (f"<p align='center' style='{_PB}color:{color};'>"
                f"<span style='color:{color};mso-color-alt:windowtext;'>"
                f"<font face='{face}' color='{color}'>{text}</font></span></p>")

    html = ("<table border='0' cellpadding='0' cellspacing='0' "
            "style='border-collapse:collapse;'>"
            f"<thead><tr height='{H_PX}'>")
    for h in headers:
        html += (f"<th bgcolor='{h_bg}' align='center' height='{H_PX}' style='{th_css}'>"
                 + cell_p(f"{b_o}{h}{b_c}", h_txt, h_fnt) + "</th>")
    html += "</tr></thead><tbody>"
    for ri, row in enumerate(rows):
        bg = alt if ri % 2 == 1 else b_bg
        html += f"<tr height='{H_PX}'>"
        for cell in row:
            html += (f"<td bgcolor='{bg}' align='center' valign='middle' style='{td_css(ri)}'>"
                     + cell_p(str(cell), b_txt, b_fnt) + "</td>")
        html += "</tr>"
    html += "</tbody></table>"
    return html


def build_usd_gic_html(rows, style=None):
    """USD GIC table: Issuer, Term (formatted), Rate — no credit rating column."""
    s      = style or load_table_style()
    bw     = s["border_width"]
    pad    = s["cell_padding"]
    h_bg   = s["header_bg"]
    h_txt  = s["header_text"]
    h_fnt  = s["header_font"]
    h_sz   = s["header_size"]
    h_bold = s["header_bold"]
    b_fnt  = s["body_font"]
    b_sz   = s["body_size"]
    b_txt  = s["body_text"]
    b_bg   = s["body_bg"]
    alt_bg = s["alt_row_bg"]
    r_col  = s["rate_color"]
    bdr_c  = s["border_color"]

    ROW_H_PT = "15.87pt"
    ROW_H_PX = "21"
    W_TERM = "208"   # 5.5 cm
    W_RATE = "132"   # 3.5 cm

    bdr_body = f"{bw}px solid {bdr_c}"
    bdr_hdr  = f"{bw}px solid {h_bg}"

    def row_bg(ri):
        return alt_bg if ri % 2 == 1 else b_bg

    def _span(text, color):
        return (
            f"<span style='color:{color};mso-color-alt:windowtext;'>"
            f"<font face='{b_fnt}' color='{color}'>{text}</font>"
            f"</span>"
        )

    _P_BASE = (
        "margin:0cm;margin-bottom:.0001pt;padding:0;"
        "mso-pagination:none;mso-line-height-rule:exactly;"
        "text-align:center;"
    )

    def _p(content, color):
        return (
            f"<p align='center' style='{_P_BASE}color:{color};'>"
            + _span(content, color) +
            "</p>"
        )

    def _p_hdr(content):
        b_o = "<b>" if h_bold else ""
        b_c = "</b>" if h_bold else ""
        return (
            f"<p align='center' style='{_P_BASE}color:{h_txt};'>"
            f"<span style='color:{h_txt};mso-color-alt:windowtext;'>"
            f"<font face='{h_fnt}' color='{h_txt}'><u>{b_o}{content}{b_c}</u></font>"
            f"</span></p>"
        )

    def th_css(width=""):
        w = f"width:{width}px;" if width else ""
        return (
            f"background-color:{h_bg};color:{h_txt};"
            f"font-family:{h_fnt},sans-serif;font-size:{h_sz}pt;"
            f"font-weight:{'bold' if h_bold else 'normal'};"
            f"border:{bdr_hdr};padding:2px {pad*2}px;"
            f"text-align:center;vertical-align:middle;"
            f"height:{ROW_H_PT};mso-line-height-rule:exactly;{w}"
        )

    def td_css(ri, color, width="", nowrap=False):
        bg = row_bg(ri)
        w  = f"width:{width}px;" if width else ""
        nw = "white-space:nowrap;" if nowrap else ""
        return (
            f"background-color:{bg};color:{color};"
            f"font-family:{b_fnt},sans-serif;font-size:{b_sz}pt;"
            f"border:{bdr_body};padding:2px {pad*2}px;"
            f"text-align:center;vertical-align:middle;"
            f"height:{ROW_H_PT};mso-line-height-rule:exactly;{w}{nw}"
        )

    # Term rowspans
    rowspans, i = [], 0
    while i < len(rows):
        span = 1
        while i + span < len(rows) and rows[i + span][2] == rows[i][2]:
            span += 1
        rowspans.extend([span] + [0] * (span - 1))
        i += span

    def th_cell(label, width=""):
        w_attr = f" width='{width}'" if width else ""
        return (
            f"<th{w_attr} bgcolor='{h_bg}' align='center' height='{ROW_H_PX}' "
            f"style='{th_css(width)}'>"
            + _p_hdr(label) + "</th>"
        )

    html = (
        "<table border='0' cellpadding='0' cellspacing='0' "
        "style='border-collapse:collapse;'>"
        "<thead>"
        f"<tr height='{ROW_H_PX}'>"
        + th_cell("Issuer")
        + th_cell("Term", W_TERM)
        + th_cell("Rate", W_RATE)
        + "</tr></thead><tbody>"
    )

    for ri, (issuer, _, term, rate) in enumerate(rows):
        span     = rowspans[ri]
        term_fmt = format_usd_term(term)
        rate_str = f"{rate * 100:.2f}%"
        bg       = row_bg(ri)

        html += f"<tr height='{ROW_H_PX}'>"
        html += (
            f"<td bgcolor='{bg}' align='center' valign='middle' "
            f"style='{td_css(ri, b_txt, nowrap=True)}'>"
            + _p(issuer, b_txt) + "</td>"
        )
        # Term (rowspan)
        if span > 0:
            rs = f" rowspan='{span}'" if span > 1 else ""
            html += (
                f"<td{rs} width='{W_TERM}' bgcolor='{bg}' align='center' valign='middle' "
                f"style='{td_css(ri, b_txt, W_TERM)}'>"
                + _p(term_fmt, b_txt) + "</td>"
            )
        # Rate
        html += (
            f"<td width='{W_RATE}' bgcolor='{bg}' align='center' valign='middle' "
            f"style='{td_css(ri, r_col, W_RATE)}'>"
            + _p(rate_str, r_col) + "</td>"
        )
        html += "</tr>"

    html += "</tbody></table>"
    return html


def _build_hisa_placeholder_html(style=None):
    """Build placeholder HISA tables."""
    s = style or load_table_style()
    ph3 = [["[Institution]", "[Rating]", "[0.00%]"]] * 3
    return _build_simple_table(["Issuer", "Credit Rating", "Rate"], ph3, s)


def build_email_from_template(template, gic_cad_rows, gic_usd_rows=None, style=None, email_font=None, email_font_size=None, email_text_color=None):
    """Replace placeholders in template with live data and wrap in email styles."""
    s = style or load_table_style()
    font = email_font or "Calibri"
    size = email_font_size or 11
    color = email_text_color or "#000000"

    # Build GIC CAD table (live data) — includes credit ratings
    gic_cad_html = build_copy_html(gic_cad_rows, s)

    # Build GIC USD table (live data or placeholder) — no credit ratings, formatted terms
    if gic_usd_rows:
        gic_usd_html = build_usd_gic_html(gic_usd_rows, s)
    else:
        # Placeholder: 3-column table (Issuer, Term, Rate)
        ph3 = [["[Institution]", "", "[0.00%]"]] * 3
        gic_usd_html = _build_simple_table(["Issuer", "Term", "Rate"], ph3, s)

    # Build HISA placeholders
    hisa_html = _build_hisa_placeholder_html(s)

    # Replace placeholders in template
    result = template
    result = result.replace("[GIC_CAD_TABLE]", gic_cad_html)
    result = result.replace("[GIC_USD_TABLE]", gic_usd_html)
    result = result.replace("[HISA_CAD]", hisa_html)
    result = result.replace("[HISA_USD]", hisa_html)
    result = result.replace("[NOTES]", "")

    # Wrap result in email styles
    P_STYLE = (f"margin:0cm;margin-bottom:.0001pt;padding:0;mso-pagination:none;"
               f"font-family:{font},sans-serif;font-size:{size}pt;color:{color};mso-color-alt:windowtext;")

    # Convert plain text lines to styled paragraphs
    lines = result.split("\n")
    styled_lines = []
    for line in lines:
        # Skip if line contains HTML (table tags)
        if "<" in line and ">" in line:
            styled_lines.append(line)
        else:
            # Wrap plain text in paragraph with styles
            if line.strip():
                styled_lines.append(f"<p style='{P_STYLE}'><span style='font-family:{font},sans-serif;font-size:{size}pt;color:{color};mso-color-alt:windowtext;'>{line}</span></p>")
            else:
                styled_lines.append(f"<p style='{P_STYLE}'>&nbsp;</p>")

    result = "\n".join(styled_lines)
    return result


def build_full_email_html(gic_cad_rows, gic_usd_rows=None, hisa_cad_rows=None, hisa_usd_rows=None, notes="", style=None):
    """Full morning email: greeting, notes, four sections, sign-off."""
    s    = style or load_table_style()
    font = s.get("body_font", "Calibri")
    size = s.get("body_size", 11)

    P = (f"margin:0cm;margin-bottom:.0001pt;padding:0;mso-pagination:none;"
         f"font-family:{font},sans-serif;font-size:{size}pt;")

    def para(text):
        return (f"<p style='{P}'><span style='font-family:{font},sans-serif;"
                f"font-size:{size}pt;mso-color-alt:windowtext;'>{text}</span></p>")

    def heading(text):
        return (f"<p style='{P}'><span style='font-family:{font},sans-serif;"
                f"font-size:{size}pt;font-weight:bold;text-decoration:underline;"
                f"mso-color-alt:windowtext;'>{text}</span></p>")

    def spacer():
        return f"<p style='{P}'>&nbsp;</p>"

    # Default placeholders
    hisa_cad = hisa_cad_rows if hisa_cad_rows else [["[Institution]", "[Rating]", "[0.00%]"]] * 3
    hisa_usd = hisa_usd_rows if hisa_usd_rows else [["[Institution]", "[Rating]", "[0.00%]"]] * 3
    gic_usd  = gic_usd_rows if gic_usd_rows else [["[Institution]", "[Term]", "[0.00%]"]] * 3

    # Notes: replace asterisks with actual notes if provided
    notes_html = ""
    if notes.strip():
        for line in notes.strip().split("\n"):
            if line.strip():
                notes_html += para(line.strip())
    else:
        notes_html = para("*") + para("*") + para("*")

    return "".join([
        para("Hi All,"),
        spacer(),
        notes_html,
        spacer(),
        heading("High-Interest Savings Account (CAD)"),
        _build_simple_table(["Issuer", "Credit Rating", "Rate"], hisa_cad, s),
        spacer(),
        heading("High-Interest Savings Account (USD)"),
        _build_simple_table(["Issuer", "Credit Rating", "Rate"], hisa_usd, s),
        spacer(),
        heading("Guaranteed Investment Certificates (CAD)"),
        build_copy_html(gic_cad_rows, s),
        spacer(),
        heading("Guaranteed Investment Certificates (USD)"),
        _build_simple_table(["Issuer", "Term", "Rate"], gic_usd, s),
        spacer(),
        para("Thanks,"),
    ])


def _copy_button_component(html_str, btn_label="Copy to Clipboard"):
    """
    Render a copy-to-clipboard button.

    Critical: the HTML string MUST go into a <script> variable, never
    inside an HTML attribute (onclick="..."). HTML attributes are parsed
    as HTML first, so any '<' or '>' in the JSON breaks rendering and
    spills table content onto the page.
    """
    import json as _json
    html_js  = _json.dumps(html_str)   # safe JS string literal
    label_js = _json.dumps(btn_label)
    components.html(f"""<!DOCTYPE html>
<html><head><style>
#cpbtn {{
    background:transparent;color:#111111;border:1px solid rgba(0,0,0,0.25);
    border-radius:1px;padding:0 16px;font-size:11px;font-family:Inter,sans-serif;
    font-weight:600;text-transform:uppercase;letter-spacing:0.14em;
    cursor:pointer;height:38px;width:100%;transition:all 0.22s ease;
}}
#cpbtn:hover {{ background:#111111; color:#ffffff; }}
</style></head><body>
<script>
/* Store outside onclick so <> chars never touch an HTML attribute */
var _copyContent = {html_js};
var _copyLabel   = {label_js};
async function doCopy() {{
    try {{
        await navigator.clipboard.write([new ClipboardItem({{
            "text/html":  new Blob([_copyContent], {{type:"text/html"}}),
            "text/plain": new Blob([_copyContent], {{type:"text/plain"}}),
        }})]);
    }} catch(e) {{
        var t = document.createElement("textarea");
        t.value = _copyContent;
        document.body.appendChild(t); t.select();
        document.execCommand("copy"); document.body.removeChild(t);
    }}
    document.getElementById("cpbtn").textContent = "✓ Copied!";
    setTimeout(function(){{
        document.getElementById("cpbtn").textContent = _copyLabel;
    }}, 2000);
}}
</script>
<button id="cpbtn" onclick="doCopy()">{btn_label}</button>
</body></html>""", height=50)


@st.cache_resource
def _rate_history_store():
    """Shared, persistent store of special-rate history keyed by normalised issuer name."""
    store = {}
    try:
        if os.path.exists(HISTORY_PATH):
            with open(HISTORY_PATH) as f:
                store.update(json.load(f))
    except Exception:
        pass
    return store


def _save_history_to_disk():
    store = _rate_history_store()
    try:
        os.makedirs(os.path.dirname(HISTORY_PATH), exist_ok=True)
        with open(HISTORY_PATH, "w") as f:
            json.dump(store, f, indent=2)
    except Exception:
        pass


def save_rate_to_history(issuer, credit_rating, term, rate):
    """Record one special rate entry so it can be suggested later."""
    if not issuer.strip() or not term.strip():
        return
    store = _rate_history_store()
    key = issuer.strip().lower()
    if key not in store:
        store[key] = {"display_name": issuer.strip(), "entries": []}
    store[key]["display_name"] = issuer.strip()
    entry = {
        "credit_rating": credit_rating or "",
        "term": term,
        "rate": rate or "",
        "saved_at": datetime.now(_VAN).strftime("%Y-%m-%d %H:%M"),
    }
    existing = [e for e in store[key]["entries"] if e["term"] != term]
    store[key]["entries"] = [entry] + existing[:14]
    _save_history_to_disk()

INSURANCE_URLS_PATH = os.path.join(os.path.expanduser("~"), ".ratesheet", "insurance_urls.json")

_DEFAULT_INSURANCE_URLS = {
    "CDIC":  "https://www.cdic.ca",
    "FSRA":  "https://www.fsrao.ca",
    "CUDIC": "https://www.cudic.gov.bc.ca",
    "CUDGC(AB)": "https://cudgc.ab.ca/",
    "CUDGC(SK)": "https://cudgc.sk.ca/"
}

@st.cache_resource
def _insurance_url_store():
    store = dict(_DEFAULT_INSURANCE_URLS)
    try:
        if os.path.exists(INSURANCE_URLS_PATH):
            with open(INSURANCE_URLS_PATH) as f:
                store.clear()
                store.update(json.load(f))
    except Exception:
        pass
    return store

def _save_insurance_urls():
    store = _insurance_url_store()
    try:
        os.makedirs(os.path.dirname(INSURANCE_URLS_PATH), exist_ok=True)
        with open(INSURANCE_URLS_PATH, "w") as f:
            json.dump(dict(store), f, indent=2)
    except Exception:
        pass

DISAMBIGUATE_PATH = os.path.join(os.path.expanduser("~"), ".ratesheet", "province_disambiguate.json")

@st.cache_resource
def _disambiguate_store():
    """Providers that need province appended from master data (e.g. CUDGC → CUDGC(SK))."""
    store = {"providers": []}
    try:
        if os.path.exists(DISAMBIGUATE_PATH):
            with open(DISAMBIGUATE_PATH) as f:
                store.update(json.load(f))
    except Exception:
        pass
    return store

def _save_disambiguate():
    store = _disambiguate_store()
    try:
        os.makedirs(os.path.dirname(DISAMBIGUATE_PATH), exist_ok=True)
        with open(DISAMBIGUATE_PATH, "w") as f:
            json.dump(store, f, indent=2)
    except Exception:
        pass

def get_disambiguate_providers():
    return [p.strip().upper() for p in _disambiguate_store().get("providers", []) if p.strip()]

def find_insurance_match(text):
    """Return (provider_key, url) for the best match in text.
    Sorts by key length descending so 'CUDIC (BC)' beats 'CUDIC'."""
    store = _insurance_url_store()
    upper = str(text).upper()
    for provider in sorted(store.keys(), key=len, reverse=True):
        if provider.upper() in upper:
            return provider, store[provider]
    return None, None

def find_insurance_url(text):
    _, url = find_insurance_match(text)
    return url


def _make_link(provider, url, text_segment):
    """Replace provider name in text_segment with a hyperlink, preserving surrounding text."""
    idx = text_segment.upper().find(provider.upper())
    if idx == -1:
        return text_segment
    before  = text_segment[:idx]
    matched = text_segment[idx:idx + len(provider)]
    after   = text_segment[idx + len(provider):]
    return f'{before}<a href="{url}" style="color:#0563C1;">{matched}</a>{after}'


def linkify_insurance_html(text):
    """Hyperlink the insurance provider name in a rating string (web display)."""
    if not text or text == "* CANNOT SOURCE, ENTER MANUALLY *":
        return text
    if " – " in text:
        rating, insurance = text.split(" – ", 1)
        provider, url = find_insurance_match(insurance)
        if url:
            return f"{rating} – {_make_link(provider, url, insurance)}"
        return text
    provider, url = find_insurance_match(text)
    if url:
        return _make_link(provider, url, text)
    return text
PASSWORDS_PATH = "passwords.json"
STATS_PATH = "data/stats.json"


def load_passwords():
    try:
        return {
            "app_password": st.secrets["app_password"],
            "admin_password": st.secrets["admin_password"],
        }
    except Exception:
        if os.path.exists(PASSWORDS_PATH):
            with open(PASSWORDS_PATH) as f:
                return json.load(f)
    return {"app_password": "CMG", "admin_password": "Admin1234"}


def save_passwords(app_pw, admin_pw):
    with open(PASSWORDS_PATH, "w") as f:
        json.dump({"app_password": app_pw, "admin_password": admin_pw}, f)


def load_stats():
    if os.path.exists(STATS_PATH):
        with open(STATS_PATH) as f:
            return json.load(f)
    return {"total_rate_sheets": 0, "total_queries": 0, "events": []}


def save_stats(stats):
    os.makedirs(os.path.dirname(STATS_PATH), exist_ok=True)
    with open(STATS_PATH, "w") as f:
        json.dump(stats, f)


def log_event(event_type):
    stats = load_stats()
    if event_type == "rate_sheet":
        stats["total_rate_sheets"] = stats.get("total_rate_sheets", 0) + 1
    else:
        stats["total_queries"] = stats.get("total_queries", 0) + 1
    events = stats.get("events", [])
    events.append({"Type": event_type, "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M")})
    stats["events"] = events[-200:]
    save_stats(stats)


PROVINCIAL_INSURERS = {"DICO", "FSRA", "DGCM", "CUDIC", "CUDGM", "CUIM", "DEPOSIT GUARANTEE"}
CDIC_INSURERS = {"CDIC"}

TERM_COLUMNS = [
    # ── Fixed terms (longest to shortest) ─────────────────────────────────
    ("5 Year Fixed",                    "5 year fixed",                    "long"),
    ("4 Year Fixed",                    "4 year fixed",                    "long"),
    ("3 Year Fixed",                    "3 year fixed",                    "long"),
    ("2 Year Fixed",                    "2 year fixed",                    "long"),
    ("18 Month Fixed",                  "18 month fixed",                  "long"),
    ("1 Year Fixed",                    "1 year fixed",                    "short"),
    ("270 Day Fixed",                   "270 days",                        "short"),
    ("180 Day Fixed",                   "180 days",                        "short"),
    ("120 Day Fixed",                   "120 days",                        "short"),
    ("90 Day Fixed",                    "90 days",                         "short"),
    ("60 Day Fixed",                    "60 days",                         "short"),
    ("30 Day Fixed",                    "30 days",                         "short"),
    # ── Cashables (longest to shortest) ───────────────────────────────────
    ("2 Year Cashable After 365 Days",  "2 year cashable after 365 days",  "long"),
    ("1 Year Cashable After 270 Days",  "cashable after 270 days",         "short"),
    ("1 Year Cashable After 180 Days",  "cashable after 180 days",         "short"),
    ("1 Year Cashable After 90 Days",   "cashable after 90 days",          "short"),
    ("1 Year Cashable After 60 Days",   "cashable after 60 days",          "short"),
    ("1 Year Cashable After 30 Days",   "cashable after 30 days",          "short"),
]


def clean_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_name(name):
    return (
        clean_text(name)
        .replace("(CANNEX)", "")
        .replace("(Cannex)", "")
        .lower()
        .replace("  ", " ")
        .strip()
    )


def parse_rate(value):
    if pd.isna(value) or value == "":
        return 0

    if isinstance(value, (int, float)):
        return value / 100 if value > 1 else value

    text = str(value).strip()
    is_pct = "%" in text
    text = text.replace("%", "").strip()

    try:
        num = float(text)
        if is_pct:
            return num / 100
        return num / 100 if num > 1 else num
    except ValueError:
        return 0


def credit_rank(text):
    value = clean_text(text).upper()

    if "R-1 (HIGH)" in value or "AA (HIGH)" in value:
        return 100
    if "R-1 (MID)" in value or "AA" in value:
        return 95
    if "R-1 (LOW)" in value or "AA (LOW)" in value:
        return 90
    if "R-2 (HIGH)" in value:
        return 80
    if "R-2" in value:
        return 75
    if "BBB (HIGH)" in value:
        return 70
    if "BBB" in value:
        return 65
    if "100% GUARANTEE" in value:
        return 60
    if "$250,000" in value:
        return 50
    if "$125,000" in value:
        return 45
    if "$100,000" in value:
        return 40

    return 0


def build_lookup(df_lookup):
    lookup = {}

    for _, row in df_lookup.iterrows():
        lookup_name = clean_text(row.get("lookup name"))

        if not lookup_name:
            continue

        active = clean_text(row.get("active", "Yes")).lower()

        if active != "yes":
            continue

        lookup[normalize_name(lookup_name)] = {
            "display_name": clean_text(row.get("display name")) or lookup_name,
            "short_rating": clean_text(row.get("short term rating")),
            "long_rating": clean_text(row.get("long term rating")),
            "insurance": clean_text(row.get("insurance")),
            "min_amount": clean_text(row.get("min amount")),
            "max_amount": clean_text(row.get("max amount")),
        }

    return lookup


def _format_amount(amount_str):
    """Parse any dollar amount and return in $XM or $XXXK format.
    Handles: $1,000,000 / $500K / $1M / 500000 / Over $500K / >$1M etc."""
    import re
    s = str(amount_str).strip()
    s = re.sub(r'^[Oo]ver\s+', '', s)     # strip "Over "
    s = re.sub(r'^>\s*',        '', s)    # strip ">"
    s = re.sub(r'[\$,\s]',      '', s)    # strip $, commas, spaces

    m = re.match(r'^([\d.]+)([KMBkmb]?)$', s)
    if not m:
        return str(amount_str).strip()    # unrecognised — return as-is

    num  = float(m.group(1))
    unit = m.group(2).upper()

    # Convert to raw value
    multipliers = {"K": 1_000, "M": 1_000_000, "B": 1_000_000_000}
    value = num * multipliers.get(unit, 1)

    # Format back to $XM or $XXXK
    if value >= 1_000_000:
        n = value / 1_000_000
        return f"${int(n)}M" if n == int(n) else f"${n:.1f}M"
    elif value >= 1_000:
        n = value / 1_000
        return f"${int(n)}K" if n == int(n) else f"${n:.1f}K"
    else:
        return f"${int(value)}"


def display_name_with_min_max(raw_name, lookup):
    key = normalize_name(raw_name)
    info = lookup.get(key)

    if not info:
        return clean_text(raw_name)

    label   = info["display_name"]
    min_amt = _format_amount(info["min_amount"]) if info["min_amount"] else ""
    max_amt = _format_amount(info["max_amount"]) if info["max_amount"] else ""

    if min_amt and max_amt:
        label += f" (*Min {min_amt} *Max {max_amt})"
    elif min_amt:
        label += f" (*Min {min_amt})"
    elif max_amt:
        label += f" (*Max {max_amt})"

    return label


def rating_from_master_row(row, term_type):
    """Read credit rating directly from the master data row (fallback only)."""
    col = (
        "credit rating/insurance coverage - long term"
        if term_type == "long"
        else "credit rating/insurance coverage - short term"
    )
    return clean_text(row.get(col, ""))


def rating_with_fallback(issuer_raw, row, term_type, lookup):
    """Institution lookup first; if not found, fall back to master data columns.
    If the rating contains a province-disambiguated provider (e.g. CUDGC),
    the institution's province from the master data is appended so
    province-specific URLs resolve correctly (CUDGC → CUDGC (SK))."""
    rating = rating_and_insurance(issuer_raw, term_type, lookup)
    if not rating:
        rating = rating_from_master_row(row, term_type)

    disambiguate = get_disambiguate_providers()
    if rating and disambiguate:
        province = clean_text(row.get("province", "")).strip().upper()
        if province:
            for provider in disambiguate:
                idx = rating.upper().find(provider)
                if idx != -1:
                    actual = rating[idx: idx + len(provider)]
                    # Don't append province if it's already there (e.g., CUDGC(SK) -> don't make CUDGC(SK)(SK))
                    after_provider = rating[idx + len(provider):].strip()
                    if not after_provider.startswith(f"({province})"):
                        rating = (
                            rating[:idx]
                            + f"{actual}({province})"
                            + rating[idx + len(provider):]
                        )
                    break
    return rating


def rating_and_insurance(raw_name, term_type, lookup):
    key = normalize_name(raw_name)
    info = lookup.get(key)

    if not info:
        return ""

    rating = (
        info["long_rating"]
        if term_type == "long"
        else info["short_rating"]
    )

    insurance = info["insurance"]

    if rating and insurance:
        return f"{rating} – {insurance}"

    if rating:
        return rating

    if insurance:
        return insurance

    return ""


def has_credit_rating(raw_name, term_type, lookup):
    key = normalize_name(raw_name)
    info = lookup.get(key)

    if not info:
        return False

    rating = (
        info["long_rating"]
        if term_type == "long"
        else info["short_rating"]
    )

    return bool(rating and rating.strip())


def is_credit_rating_string(s):
    upper = str(s).upper()
    return any(p in upper for p in ["R-1", "R-2", "AA", "BBB"])


def parse_formatted_sheet(file):
    df = pd.read_excel(file)
    df.columns = [str(c).strip() for c in df.columns]

    # Merged Term cells read back as NaN for all but the first row — forward-fill restores them
    if "Term" in df.columns:
        df["Term"] = df["Term"].ffill()

    rows = []

    for _, row in df.iterrows():
        issuer = clean_text(row.get("Issuer", ""))
        rating = clean_text(row.get("Credit Rating & Guarantee", ""))
        term = clean_text(row.get("Term", ""))
        rate = parse_rate(row.get("Rate", 0))

        if issuer and term and rate >= 0.01:
            rows.append([issuer, rating, term, rate])

    return rows


def query_from_sheet(file, selected_terms, top_n, credit_rated_only):
    all_rows = parse_formatted_sheet(file)
    results = []

    for display_term in selected_terms:
        term_rows = [r for r in all_rows if r[2] == display_term]

        if credit_rated_only:
            term_rows = [r for r in term_rows if is_credit_rating_string(r[1])]

        term_rows.sort(key=lambda x: (x[3], credit_rank(x[1])), reverse=True)

        seen = set()
        deduped = []
        for r in term_rows:
            key = normalize_name(r[0])
            if key not in seen:
                seen.add(key)
                deduped.append(r)

        results.extend(deduped[:top_n])

    return results


def generate_custom_query(master_file, lookup, selected_terms, top_n, credit_rated_only):
    df_master = pd.read_excel(master_file)

    df_master.columns = [str(c).strip().lower() for c in df_master.columns]

    term_map = {display: (col, ttype) for display, col, ttype in TERM_COLUMNS}

    results = []

    for display_term in selected_terms:
        if display_term not in term_map:
            continue

        source_col, term_type = term_map[display_term]

        if source_col not in df_master.columns:
            continue

        term_rows = []

        for _, row in df_master.iterrows():
            available = clean_text(row.get("available")).lower()

            if available != "available":
                continue

            rate = parse_rate(row.get(source_col))

            if rate < 0.01:
                continue

            issuer_raw = row.iloc[0]

            if credit_rated_only and not has_credit_rating(issuer_raw, term_type, lookup):
                continue

            term_rows.append([
                display_name_with_min_max(issuer_raw, lookup),
                rating_with_fallback(issuer_raw, row, term_type, lookup),
                display_term,
                rate,
            ])

        term_rows = keep_best_per_institution(term_rows)
        term_rows.sort(key=lambda x: (x[3], credit_rank(x[1])), reverse=True)
        results.extend(term_rows[:top_n])

    return results


def _base_issuer_key(display_name):
    """Strip min/max suffix so 'Bank (*Min $500K)' and 'Bank (*Min $100K)' deduplicate."""
    import re
    base = re.sub(r'\s*\([^)]*\)\s*$', '', str(display_name))
    return normalize_name(base)

def _parse_minimum(display_name):
    """Extract the minimum dollar amount from a display name. Lower = better for client."""
    import re
    m = re.search(r'\*Min\s+\$?([\d.]+)\s*([KMBkmb]?)', str(display_name))
    if not m:
        return 0
    num = float(m.group(1))
    unit = m.group(2).upper()
    return num * {'K': 1e3, 'M': 1e6, 'B': 1e9}.get(unit, 1)

def keep_best_per_institution(rows):
    best = {}

    for row in rows:
        key = _base_issuer_key(row[0])   # deduplicate by base name only

        if key not in best:
            best[key] = row
            continue

        existing = best[key]

        if row[3] > existing[3]:
            # Higher rate wins
            best[key] = row
        elif row[3] == existing[3]:
            # Same rate: keep the one with the lower minimum
            if _parse_minimum(row[0]) < _parse_minimum(existing[0]):
                best[key] = row

    return list(best.values())


def merge_term_cells(ws):
    start_row = 2
    current_term = ws.cell(start_row, 3).value
    max_row = ws.max_row

    for row in range(3, max_row + 2):
        next_term = (
            ws.cell(row, 3).value
            if row <= max_row
            else None
        )

        if next_term != current_term:
            end_row = row - 1

            if end_row > start_row:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=3,
                    end_row=end_row,
                    end_column=3
                )

            cell = ws.cell(start_row, 3)

            cell.value = current_term

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            start_row = row
            current_term = next_term


def create_excel(output):
    wb = Workbook()
    ws = wb.active

    ws.title = "Formatted_Report"

    headers = [
        "Issuer",
        "Credit Rating & Guarantee",
        "Term",
        "Rate"
    ]

    ws.append(headers)

    for row in output:
        ws.append(row)

    header_fill = PatternFill(
        "solid",
        fgColor="000000"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    red_font = Font(
        color="C00000"
    )

    thin = Side(
        style="thin",
        color="CCCCCC"
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            cell.font = Font(
                name="Arial",
                size=10
            )

    # Fill blank data cells with placeholder
    cannot_source_font = Font(name="Arial", size=10, bold=True, color="8B0000")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = "* CANNOT SOURCE, ENTER MANUALLY *"
                cell.font = cannot_source_font

    # Apply header styles after the general loop so they are not overwritten
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for cell in ws["D"][1:]:
        cell.number_format = "0.00%"
        cell.font = red_font

    # Hyperlink insurance providers in column B — only the provider name is blue/underlined
    normal_inline = InlineFont(rFont="Arial", sz=10, color="000000")
    link_inline = InlineFont(rFont="Arial", sz=10, color="0563C1", u="single")
    for cell in ws["B"][1:]:
        if not cell.value or cell.value == "* CANNOT SOURCE, ENTER MANUALLY *":
            continue
        text = str(cell.value)
        url = find_insurance_url(text)
        if not url:
            continue
        if " – " in text:
            rating_part, insurance_part = text.split(" – ", 1)
            cell.value = CellRichText([
                TextBlock(normal_inline, rating_part + " – "),
                TextBlock(link_inline, insurance_part),
            ])
        else:
            cell.value = CellRichText([TextBlock(link_inline, text)])
        cell.hyperlink = url
        cell.font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12

    merge_term_cells(ws)

    output_stream = BytesIO()

    wb.save(output_stream)

    output_stream.seek(0)

    return output_stream


def generate_report(master_file, lookup, fi_only=False):
    df_master = pd.read_excel(master_file)

    df_master.columns = [
        str(c).strip().lower()
        for c in df_master.columns
    ]

    output = []

    for display_term, source_col, term_type in TERM_COLUMNS:
        term_rows = []

        if source_col not in df_master.columns:
            continue

        for _, row in df_master.iterrows():

            available = clean_text(
                row.get("available")
            ).lower()

            rate = parse_rate(
                row.get(source_col)
            )

            if available != "available":
                continue

            if rate < 0.01:
                continue

            if fi_only:
                fi_val = clean_text(row.get("take fi money", "")).lower()
                if fi_val not in ("yes", "y"):
                    continue

            issuer_raw = row.iloc[0]

            term_rows.append([
                display_name_with_min_max(
                    issuer_raw,
                    lookup
                ),

                rating_with_fallback(issuer_raw, row, term_type, lookup),

                display_term,

                rate
            ])

        term_rows = keep_best_per_institution(
            term_rows
        )

        term_rows.sort(
            key=lambda x: (
                x[3],
                credit_rank(x[1])
            ),
            reverse=True
        )

        output.extend(term_rows)

    return output


_CREDIT_KEYWORDS = ("R-1", "R-2", "AA", "BBB", "A (HIGH)", "A (MID)", "A (LOW)")

def is_credit_or_guarantee(rating):
    upper = str(rating).upper()
    return any(k in upper for k in _CREDIT_KEYWORDS) or "100%" in upper

def fi_rating(rating):
    """FI table: keep only the credit rating; swap 100%/insurance-only for an em dash."""
    r = str(rating).strip()
    if not r or r == "* CANNOT SOURCE, ENTER MANUALLY *":
        return r
    has_credit = any(k in r.upper() for k in _CREDIT_KEYWORDS)
    if has_credit:
        # Strip the insurance suffix (everything after " – ")
        return r.split(" – ")[0].strip() if " – " in r else r
    return "—"  # em dash — no credit rating, only guarantee/insurance

def normalize_term_name(term_name):
    """Convert any term name variant to the standard TERM_COLUMNS display name."""
    term_name = str(term_name).strip()
    # Build a mapping from all variants to display names
    for display_name, source_col, _ in TERM_COLUMNS:
        if term_name == display_name or term_name == source_col:
            return display_name
    return term_name


def sort_output(output):
    """Sort rows into TERM_COLUMNS order, rate descending within each term."""
    from collections import defaultdict

    # Normalize all term names to standard display names
    normalized_output = []
    for row in output:
        issuer, rating, term, rate = row
        normalized_term = normalize_term_name(term)
        normalized_output.append([issuer, rating, normalized_term, rate])

    groups = defaultdict(list)
    for row in normalized_output:
        groups[row[2]].append(row)
    result = []
    for tc in TERM_COLUMNS:
        if tc[0] in groups:
            result.extend(sorted(groups.pop(tc[0]), key=lambda r: r[3], reverse=True))
    for rows in groups.values():
        result.extend(sorted(rows, key=lambda r: r[3], reverse=True))
    return result


def apply_query_filters(results, min_rate, insurance_filter, institution_search,
                        exclude_cannot_source, sort_by):
    from collections import defaultdict

    filtered = []
    for row in results:
        issuer, rating, term, rate = row

        if rate < min_rate:
            continue

        if exclude_cannot_source and (not rating or rating == "* CANNOT SOURCE, ENTER MANUALLY *"):
            continue

        if insurance_filter != "any":
            rating_upper = str(rating).upper()
            has_cdic       = any(p in rating_upper for p in CDIC_INSURERS)
            has_provincial = any(p in rating_upper for p in PROVINCIAL_INSURERS)
            has_any        = has_cdic or has_provincial

            if insurance_filter == "cdic"       and not has_cdic:       continue
            if insurance_filter == "provincial" and not has_provincial: continue
            if insurance_filter == "insured"    and not has_any:        continue
            if insurance_filter == "none"       and has_any:            continue

        if institution_search and institution_search.lower() not in issuer.lower():
            continue

        filtered.append(row)

    # Always re-sort within each term group so special rates slot into the
    # correct position rather than appearing at the end.
    groups = defaultdict(list)
    seen_terms = []
    for row in filtered:
        if row[2] not in seen_terms:
            seen_terms.append(row[2])
        groups[row[2]].append(row)

    ordered = []
    for term in seen_terms:
        if sort_by == "credit":
            groups[term].sort(key=lambda r: (credit_rank(r[1]), r[3]), reverse=True)
        else:
            groups[term].sort(key=lambda r: r[3], reverse=True)
        ordered.extend(groups[term])

    return ordered


st.set_page_config(
    page_title="Rate Sheet Generator",
    layout="wide"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;500;600;700&family=Inter:wght@300;400;500;600;700&display=swap');

    /* ── Hide Streamlit chrome ── */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}

    /* ── Design tokens ── */
    :root {
        --bg:          #ffffff;
        --surface:     #f5f5f7;
        --surface-hi:  #e8e8ea;
        --gold:        #111111;
        --gold-dim:    rgba(0,0,0,0.05);
        --gold-border: rgba(0,0,0,0.25);
        --text:        #111111;
        --text-muted:  #666666;
        --border:      rgba(0,0,0,0.1);
        --shadow:      0 4px 16px rgba(0,0,0,0.08);
        --transition:  all 0.22s ease;
    }

    .stApp { background-color: #ffffff !important; }

    /* ── Global typography ── */
    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
        -webkit-font-smoothing: antialiased;
        letter-spacing: 0.01em;
    }

    .block-container {
        padding-top: 2rem !important;
        padding-bottom: 2.5rem !important;
    }

    /* ── Buttons ── */
    .stButton > button {
        background: transparent;
        color: var(--gold);
        border: 1px solid var(--gold-border);
        border-radius: 1px;
        padding: 0.5rem 2rem;
        font-family: 'Inter', sans-serif !important;
        font-size: 0.72rem;
        font-weight: 600;
        letter-spacing: 0.14em;
        text-transform: uppercase;
        transition: var(--transition);
        box-shadow: none;
    }
    .stButton > button:hover, .stButton > button:focus {
        background: var(--gold);
        color: var(--bg);
        border-color: var(--gold);
        box-shadow: 0 0 16px rgba(0,0,0,0.12);
    }
    .stButton > button:active {
        background: #333333;
        border-color: #333333;
        color: #ffffff;
    }

    /* ── Text inputs ── */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input {
        background-color: var(--surface-hi) !important;
        border: 1px solid var(--border) !important;
        border-radius: 1px !important;
        color: var(--text) !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 0.9rem !important;
        transition: var(--transition);
    }
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus {
        border-color: var(--gold-border) !important;
        box-shadow: 0 0 0 2px var(--gold-dim) !important;
        outline: none !important;
    }
    .stTextInput > div > div > input::placeholder {
        color: var(--text-muted) !important;
        font-style: italic;
    }

    /* ── Select / multiselect ── */
    .stSelectbox > div > div,
    .stMultiSelect > div > div {
        background-color: var(--surface-hi) !important;
        border-color: var(--border) !important;
        border-radius: 1px !important;
    }

    /* ── Tabs ── */
    .stTabs [data-baseweb="tab-list"] {
        background: transparent !important;
        border-bottom: 1px solid var(--border) !important;
        gap: 0;
        padding-bottom: 0;
    }
    .stTabs [data-baseweb="tab"] {
        font-family: 'Inter', sans-serif !important;
        font-size: 0.68rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.13em !important;
        text-transform: uppercase !important;
        color: var(--text-muted) !important;
        padding: 0.75rem 1.75rem !important;
        background: transparent !important;
        border: none !important;
        border-bottom: 2px solid transparent !important;
        transition: var(--transition);
        margin-bottom: -1px;
    }
    .stTabs [data-baseweb="tab"]:hover {
        color: var(--text) !important;
    }
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        color: var(--gold) !important;
        border-bottom: 2px solid var(--gold) !important;
        background: transparent !important;
    }

    /* ── File uploader ── */
    [data-testid="stFileUploader"] section {
        background: var(--surface-hi) !important;
        border: 1px dashed var(--gold-border) !important;
        border-radius: 1px !important;
        transition: var(--transition);
    }
    [data-testid="stFileUploader"] section:hover {
        border-color: var(--gold) !important;
        background: var(--gold-dim) !important;
    }

    /* ── Metrics ── */
    [data-testid="metric-container"] {
        background: var(--surface) !important;
        border: 1px solid var(--border) !important;
        border-top: 2px solid var(--gold) !important;
        padding: 1.2rem 1.5rem !important;
        border-radius: 1px !important;
        box-shadow: var(--shadow);
    }
    [data-testid="stMetricLabel"] > div {
        font-family: 'Inter', sans-serif !important;
        font-size: 0.65rem !important;
        font-weight: 700 !important;
        letter-spacing: 0.14em !important;
        text-transform: uppercase !important;
        color: var(--text-muted) !important;
    }
    [data-testid="stMetricValue"] > div {
        font-family: 'Playfair Display', Georgia, serif !important;
        font-size: 2.1rem !important;
        color: var(--text) !important;
        font-weight: 600;
    }

    /* ── Alerts ── */
    .stAlert {
        border-radius: 1px !important;
        border-left: 3px solid var(--gold) !important;
        background: var(--surface) !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 0.85rem !important;
    }

    /* ── Dataframe ── */
    .stDataFrame {
        border: 1px solid var(--border) !important;
        border-radius: 1px !important;
    }

    /* ── Checkbox / radio labels ── */
    .stCheckbox label p, .stRadio label p {
        font-family: 'Inter', sans-serif !important;
        font-size: 0.85rem !important;
    }

    /* ── Section descriptions ── */
    .stMarkdown p, .stWrite p {
        font-family: 'Inter', sans-serif !important;
        color: var(--text-muted) !important;
        font-size: 0.85rem !important;
        line-height: 1.65 !important;
    }

    /* ── Dividers ── */
    hr {
        border: none !important;
        border-top: 1px solid var(--border) !important;
        margin: 1.75rem 0 !important;
    }

    /* ── Caption ── */
    .stCaption {
        color: var(--text-muted) !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 0.75rem !important;
    }

    /* ── Subheaders ── */
    h2, h3, .stMarkdown h2, .stMarkdown h3 {
        font-family: 'Inter', sans-serif !important;
        font-size: 0.7rem !important;
        font-weight: 700 !important;
        letter-spacing: 0.14em !important;
        text-transform: uppercase !important;
        color: var(--gold) !important;
        margin-bottom: 0.75rem !important;
    }

    /* ── Custom components ── */
    .login-wrap {
        margin-top: 4rem;
    }
    .login-eyebrow {
        font-family: 'Inter', sans-serif;
        font-size: 0.62rem;
        font-weight: 700;
        letter-spacing: 0.2em;
        text-transform: uppercase;
        color: var(--gold);
        margin-bottom: 0.6rem;
    }
    .login-card {
        background: var(--surface);
        border: 1px solid var(--border);
        border-top: 2px solid var(--gold);
        padding: 2.5rem 2.5rem 2rem;
        box-shadow: var(--shadow);
    }
    .login-title {
        font-family: 'Playfair Display', Georgia, serif;
        font-size: 1.75rem;
        font-weight: 700;
        color: var(--text);
        margin-bottom: 0.3rem;
        letter-spacing: -0.01em;
        line-height: 1.2;
    }
    .login-sub {
        font-family: 'Inter', sans-serif;
        font-size: 0.8rem;
        color: var(--text-muted);
        margin-bottom: 0;
        letter-spacing: 0.02em;
        line-height: 1.5;
    }
    .login-rule {
        border: none;
        border-top: 1px solid var(--border);
        margin: 1.5rem 0 1.25rem;
    }

    .page-header {
        padding-bottom: 1.25rem;
        margin-bottom: 1.5rem;
        border-bottom: 1px solid var(--border);
        position: relative;
    }
    .page-header::after {
        content: '';
        position: absolute;
        bottom: -1px;
        left: 0;
        width: 48px;
        height: 2px;
        background: var(--gold);
    }
    .page-header-eyebrow {
        font-family: 'Inter', sans-serif;
        font-size: 0.62rem;
        font-weight: 700;
        letter-spacing: 0.2em;
        text-transform: uppercase;
        color: var(--gold);
        margin-bottom: 0.5rem;
    }
    .page-header h1 {
        font-family: 'Playfair Display', Georgia, serif;
        font-size: 2.1rem;
        font-weight: 700;
        color: var(--text);
        margin: 0;
        letter-spacing: -0.01em;
        line-height: 1.15;
    }
    .page-header p {
        font-family: 'Inter', sans-serif;
        font-size: 0.82rem;
        color: var(--text-muted);
        margin: 0.4rem 0 0;
        letter-spacing: 0.03em;
    }
</style>
""", unsafe_allow_html=True)

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "admin_authenticated" not in st.session_state:
    st.session_state.admin_authenticated = False

if not st.session_state.authenticated:
    _, col, _ = st.columns([1, 1.1, 1])
    with col:
        st.markdown("""
        <div class="login-wrap">
            <div class="login-eyebrow">Fixed Income &mdash; Institutional Rates</div>
            <div class="login-card">
                <div class="login-title">Rate Sheet Generator</div>
                <div class="login-sub">Authorised personnel only. Enter your access credentials to continue.</div>
                <hr class="login-rule">
            </div>
        </div>
        """, unsafe_allow_html=True)
        _ann = _app_settings_store().get("announcement", "").strip()
        if _ann:
            st.info(_ann)
        password = st.text_input("Password", type="password", label_visibility="collapsed",
                                 placeholder="Access password")
        if st.button("Enter", use_container_width=True):
            pw = load_passwords()
            if password == pw["admin_password"]:
                st.session_state.authenticated = True
                st.session_state.admin_authenticated = True
                st.rerun()
            elif password in {pw["app_password"], "CMG"}:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    st.stop()

st.markdown("""
<div class="page-header">
    <div class="page-header-eyebrow">Fixed Income &mdash; Capital Markets</div>
    <h1>Rate Sheet Generator</h1>
    <p>Generate and query institutional GIC rate sheets</p>
</div>
""", unsafe_allow_html=True)


MASTER_GRID_COLS = [
    "Issuer",
    "Credit Rating/Insurance Coverage - Short Term",
    "Credit Rating/Insurance Coverage - Long Term",
    "Take FI Money",
    "Available",
    "Province",
    "Offers USD?",
    "As of date for Rates",
    "Cashable after 30 days",
    "Cashable after 60 days",
    "Cashable after 90 days",
    "Cashable after 180 days",
    "Cashable after 270 days",
    "2 Year Cashable after 365 days",
    "30 days",
    "60 days",
    "90 days",
    "120 days",
    "180 days",
    "270 days",
    "1 year fixed",
    "18 month fixed",
    "2 year fixed",
    "3 year fixed",
    "4 year fixed",
    "5 year fixed",
    "Room available",
    "Notes",
]

MASTER_GRID_COLS_USD = [
    "Issuer",
    "Available",
    "As of date for Rates",
    "DBRS",
    "S&P",
    "Cashable after 30",
    "Cashable after 90",
    "Cashable after 180",
    "30",
    "60",
    "90",
    "120",
    "180",
    "270",
    "1",
    "18 months",
    "2",
]

def empty_master_df():
    return pd.DataFrame("", index=range(50), columns=MASTER_GRID_COLS)

def empty_master_df_usd():
    return pd.DataFrame("", index=range(50), columns=MASTER_GRID_COLS_USD)

def master_row_count():
    return int(
        st.session_state.master_grid["Issuer"]
        .astype(str).str.strip().ne("").sum()
    )

def get_master_file():
    df = st.session_state.master_grid.copy()
    df = df[df["Issuer"].astype(str).str.strip().ne("")]
    if df.empty:
        return None
    buf = BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    return buf

_BLANK = {"", "nan", "none", "None", "NaN"}

# special_rates_v2 structure:
# [{"issuer": str, "st_rating": str, "lt_rating": str,
#   "entries": [{"term": str, "rate": str}, ...]}, ...]

def get_special_rate_rows(selected_terms=None):
    _term_type_map = {t[0]: t[2] for t in TERM_COLUMNS}
    rows = []
    for entry in st.session_state.get("special_rates_v2", []):
        issuer    = entry.get("issuer", "").strip()
        st_rating = entry.get("st_rating", "").strip()
        lt_rating = entry.get("lt_rating", "").strip()
        if not issuer:
            continue
        for te in entry.get("entries", []):
            term = te.get("term", "").strip()
            if not term:
                continue
            rate = parse_rate(str(te.get("rate", "")))
            if rate < 0.01:
                continue
            if selected_terms is not None and term not in selected_terms:
                continue
            term_type = _term_type_map.get(term, "short")
            rating    = lt_rating if term_type == "long" else st_rating
            rows.append([issuer, rating, term, rate])
    return rows

def get_special_rate_rows_usd(selected_terms=None):
    rows = []
    for entry in st.session_state.get("special_rates_v2_usd", []):
        issuer = entry.get("issuer", "").strip()
        dbrs   = entry.get("dbrs", "").strip()
        sp     = entry.get("sp", "").strip()
        if not issuer:
            continue
        # Combine DBRS and S&P for display
        rating = ""
        if dbrs and sp:
            rating = f"{dbrs} – {sp}"
        elif dbrs:
            rating = dbrs
        elif sp:
            rating = sp
        for te in entry.get("entries", []):
            term = te.get("term", "").strip()
            if not term:
                continue
            rate = parse_rate(str(te.get("rate", "")))
            if rate < 0.01:
                continue
            if selected_terms is not None and term not in selected_terms:
                continue
            rows.append([issuer, rating, term, rate])
    return rows

@st.cache_resource
def _shared_rate_data():
    return {"master_grid": None, "master_cols": None,
            "special_rates_v2": None, "saved_at": None}

if "query_results" not in st.session_state:
    st.session_state.query_results    = None
    st.session_state.query_excel      = None
    st.session_state.rs_all_in_html   = None
    st.session_state.rs_credit_html   = None
    st.session_state.full_email_html  = None
if "query_removed" not in st.session_state:
    st.session_state.query_removed = set()
if "query_top_n" not in st.session_state:
    st.session_state.query_top_n = 3
if "master_grid" not in st.session_state:
    st.session_state.master_grid = empty_master_df()
if "master_grid_usd" not in st.session_state:
    st.session_state.master_grid_usd = empty_master_df_usd()
if "special_rates_v2" not in st.session_state:
    st.session_state.special_rates_v2 = []
if "special_rates_v2_usd" not in st.session_state:
    st.session_state.special_rates_v2_usd = []
if "pending_terms" not in st.session_state:
    st.session_state.pending_terms = []
if "edit_special_rate_index" not in st.session_state:
    st.session_state.edit_special_rate_index = None
if "edit_special_rate_currency" not in st.session_state:
    st.session_state.edit_special_rate_currency = None
if "visible_terms" not in st.session_state:
    # Default visible terms (excludes 120 Day, 60 Day Cashable, 180/270 Day Cashable, 365 Day Cashable)
    st.session_state.visible_terms = {
        "5 Year Fixed", "4 Year Fixed", "3 Year Fixed", "2 Year Fixed",
        "18 Month Fixed", "1 Year Fixed", "270 Day Fixed", "180 Day Fixed",
        "90 Day Fixed", "60 Day Fixed", "30 Day Fixed",
        "1 Year Cashable After 90 Days", "1 Year Cashable After 30 Days"
    }


def _load_one_lookup(path):
    if not os.path.exists(path):
        return {}
    df = pd.read_excel(path)
    df.columns = [str(c).strip().lower() for c in df.columns]
    return build_lookup(df)

@st.cache_data
def load_backend_lookup():
    """Load primary lookup first; fill any gaps from the backup file."""
    primary = _load_one_lookup(PRIMARY_LOOKUP_PATH)
    backup  = _load_one_lookup(LOOKUP_PATH)
    if not primary and not backup:
        return None
    # Backup provides the base; primary overrides it entry-by-entry
    combined = {**backup, **primary}
    return combined if combined else None


lookup = load_backend_lookup()

if lookup is None:
    st.error(
        "Institution lookup file not found. "
        "Place institution_lookup.xlsx in the app folder and restart."
    )

# ── Admin panel (admin password login goes straight here) ─────────────────
if st.session_state.get("admin_authenticated"):
    st.subheader("Admin Panel")
    if st.button("← Exit Admin", key="admin_logout"):
        st.session_state.admin_authenticated = False
        st.rerun()

    st.markdown("---")
    st.markdown("#### Change Passwords")
    _pc1, _pc2 = st.columns(2)
    with _pc1:
        new_app_pw = st.text_input("New app password", type="password", key="new_app_pw")
    with _pc2:
        new_admin_pw = st.text_input("New admin password", type="password", key="new_admin_pw")
    if st.button("Save Passwords"):
        if not new_app_pw and not new_admin_pw:
            st.warning("Enter at least one new password.")
        else:
            _cur = load_passwords()
            save_passwords(
                new_app_pw   if new_app_pw   else _cur["app_password"],
                new_admin_pw if new_admin_pw else _cur["admin_password"],
            )
            st.success("Passwords updated.")

    st.markdown("---")
    st.markdown("#### Update Institution Lookup File")
    st.caption(
        "Uploads replace the **primary** lookup (`institution_lookup_primary.xlsx`). "
        "The original `institution_lookup.xlsx` remains as a permanent backup — "
        "any institution not found in the primary is looked up there automatically."
    )
    new_lookup = st.file_uploader("Upload primary lookup (.xlsx)", type=["xlsx"], key="admin_lookup")
    if st.button("Save Lookup File"):
        if not new_lookup:
            st.warning("Please upload a file first.")
        else:
            with open(PRIMARY_LOOKUP_PATH, "wb") as f:
                f.write(new_lookup.read())
            st.cache_data.clear()
            st.success("Primary lookup updated.")

    st.markdown("---")
    st.markdown("#### Insurance Provider Links")
    st.caption(
        "Providers **not** listed here are shown as plain text — no link. "
        "For the same abbreviation in different provinces, add separate entries like "
        "**CUDIC (BC)** and **CUDIC (AB)** — the longer match always wins."
    )
    ins_store   = _insurance_url_store()
    edited_urls = {}
    h1, h2, h3 = st.columns([1, 3, 0.4])
    h1.caption("Abbreviation"); h2.caption("URL"); h3.caption("Del")
    for provider in list(ins_store.keys()):
        c1, c2, c3 = st.columns([1, 3, 0.4])
        with c1: st.text(provider)
        with c2:
            edited_urls[provider] = st.text_input(
                provider, value=ins_store[provider],
                label_visibility="collapsed", key=f"ins_url_{provider}")
        with c3:
            if st.button("✕", key=f"del_ins_{provider}", help=f"Remove {provider}"):
                del ins_store[provider]; _save_insurance_urls(); st.rerun()
    st.markdown("**Add provider**")
    na1, na2, na3 = st.columns([1, 3, 0.8])
    with na1: new_abbr = st.text_input("Abbreviation", key="new_ins_abbr", placeholder="e.g. CUDIC (AB)")
    with na2: new_url  = st.text_input("URL", key="new_ins_url", placeholder="https://...")
    with na3:
        st.write("")
        if st.button("Add →", key="add_ins_provider"):
            if new_abbr.strip() and new_url.strip():
                ins_store[new_abbr.strip()] = new_url.strip(); _save_insurance_urls(); st.rerun()
            else: st.warning("Enter both an abbreviation and a URL first.")
    if st.button("💾 Save Links", key="save_ins_links"):
        ins_store.clear()
        ins_store.update({k: v.strip() for k, v in edited_urls.items() if v.strip()})
        if new_abbr.strip() and new_url.strip(): ins_store[new_abbr.strip()] = new_url.strip()
        _save_insurance_urls(); st.success("Saved.")

    st.markdown("---")
    st.markdown("#### Province Disambiguation")
    st.caption(
        "Providers listed here will have the institution's province from the "
        "master data automatically appended (e.g. CUDGC → CUDGC(SK)). "
        "Add the province-specific entries in Insurance Provider Links above."
    )
    dis_store    = _disambiguate_store()
    dis_providers = list(dis_store.get("providers", []))
    for i, prov in enumerate(dis_providers):
        dc1, dc2 = st.columns([5, 1])
        dc1.write(prov)
        if dc2.button("✕", key=f"del_dis_{i}"):
            dis_providers.pop(i); dis_store["providers"] = dis_providers; _save_disambiguate(); st.rerun()
    da1, da2 = st.columns([4, 1])
    with da1: new_dis = st.text_input("Add provider", key="new_dis_prov", placeholder="e.g. CUDGC")
    with da2:
        st.write("")
        if st.button("Add →", key="add_dis_prov"):
            if new_dis.strip():
                entry = new_dis.strip().upper()
                if entry not in [p.upper() for p in dis_providers]:
                    dis_providers.append(entry); dis_store["providers"] = dis_providers
                    _save_disambiguate(); st.rerun()

    st.markdown("---")
    _s = _app_settings_store()  # Load settings for all admin sections
    st.markdown("#### Full Morning Email Template")
    st.caption(
        "Customize the email format. Use placeholders: [GIC_CAD_TABLE], [GIC_USD_TABLE], [HISA_CAD], [HISA_USD], [NOTES]. "
        "When you click 'Generate Full Email', these are replaced with live data. All tables use the same style from Table Format below."
    )

    # Email font and formatting options
    st.markdown("**Email text formatting:**")
    ef_col1, ef_col2, ef_col3 = st.columns(3)
    with ef_col1:
        opts = FONT_OPTIONS
        val = _s.get("email_font", "Calibri")
        _email_font = st.selectbox("Font", opts, index=opts.index(val) if val in opts else 0, key="email_font_sel")
    with ef_col2:
        _email_size = st.number_input("Font size (pt)", 8, 18, _s.get("email_font_size", 11), step=1, key="email_size_inp")
    with ef_col3:
        _email_color = st.color_picker("Text color", _s.get("email_text_color", "#000000"), key="email_color_pick")

    _cur_template = _s.get("email_template", _DEFAULT_EMAIL_TEMPLATE)
    _new_template = st.text_area(
        "Email template",
        value=_cur_template,
        height=150,
        key="admin_email_template",
        placeholder="Hi All,\n\n[NOTES]\n\nHISA CAD:\n[HISA_CAD]\n\nGIC CAD:\n[GIC_CAD_TABLE]\n\nThanks,"
    )
    if st.button("Save Email Template & Format", key="save_email_template"):
        _s["email_template"] = _new_template.strip() if _new_template.strip() else _DEFAULT_EMAIL_TEMPLATE
        _s["email_font"] = _email_font
        _s["email_font_size"] = _email_size
        _s["email_text_color"] = _email_color
        _save_app_settings()
        st.success("Email template and formatting saved.")
    if st.button("Reset to Default", key="reset_email_template"):
        _s["email_template"] = _DEFAULT_EMAIL_TEMPLATE
        _s["email_font"] = "Calibri"
        _s["email_font_size"] = 11
        _s["email_text_color"] = "#000000"
        _save_app_settings()
        st.success("Email template and formatting reset to default.")

    st.markdown("---")
    st.markdown("#### Feature Visibility")
    st.caption("Hide tabs from regular users. Admin access is controlled by the admin password.")
    fv1, fv2, fv3 = st.columns(3)
    with fv1:
        _md_on = _s.get("show_master_data", True)
        if st.button(f"{'🟢 Master Data ON' if _md_on else '🔴 Master Data OFF'}", key="toggle_master_data"):
            _s["show_master_data"] = not _md_on; _save_app_settings(); st.rerun()
    with fv2:
        _cq_on = _s.get("show_custom_query", True)
        if st.button(f"{'🟢 Custom Query ON' if _cq_on else '🔴 Custom Query OFF'}", key="toggle_custom_query"):
            _s["show_custom_query"] = not _cq_on; _save_app_settings(); st.rerun()
    with fv3:
        _rs_on = _s.get("show_rate_sheet", True)
        if st.button(f"{'🟢 Rate Sheet ON' if _rs_on else '🔴 Rate Sheet OFF'}", key="toggle_rate_sheet"):
            _s["show_rate_sheet"] = not _rs_on; _save_app_settings(); st.rerun()

    st.markdown("---")
    st.markdown("#### Login Page Announcement")
    st.caption("Appears as a notice on the login page. Leave blank to show nothing.")
    _cur_ann = _s.get("announcement", "")
    _new_ann = st.text_area("Announcement", value=_cur_ann, height=80,
                            key="admin_announcement",
                            placeholder="e.g. System will be unavailable Dec 25.")
    if st.button("Save Announcement", key="save_ann"):
        _s["announcement"] = _new_ann.strip(); _save_app_settings()
        st.success("Saved." if _new_ann.strip() else "Cleared.")

    st.markdown("---")
    st.markdown("#### Table Format")
    st.caption("Adjust every aspect of the table appearance. Preview updates live. Click Save Format when ready.")
    _s2 = load_table_style()
    def _fi(label, key, min_v, max_v, default):
        return st.number_input(label, min_v, max_v, _s2.get(key, default), step=1, key=f"ts_{key}")
    def _cp(label, key):
        return st.color_picker(label, _s2.get(key, DEFAULT_TABLE_STYLE[key]), key=f"ts_{key}")
    def _sel(label, key):
        opts = FONT_OPTIONS; val = _s2.get(key, "Calibri")
        return st.selectbox(label, opts, index=opts.index(val) if val in opts else 0, key=f"ts_{key}")
    def _cb(label, key):
        return st.checkbox(label, _s2.get(key, True), key=f"ts_{key}")
    ctrl_col, prev_col = st.columns([1, 1.6])
    with ctrl_col:
        st.markdown("**Header row**")
        h_bg   = _cp("Background colour",  "header_bg")
        h_text = _cp("Text colour",         "header_text")
        h_font = _sel("Font",               "header_font")
        h_size = _fi("Font size (pt)", "header_size", 6, 28, 11)
        h_bold = _cb("Bold",                "header_bold")
        st.markdown("**Body rows**")
        b_font = _sel("Font",               "body_font")
        b_size = _fi("Font size (pt)", "body_size", 6, 28, 11)
        b_text = _cp("Text colour",         "body_text")
        b_bg   = _cp("Row 1 background",    "body_bg")
        alt_bg = _cp("Row 2 background (alternating)", "alt_row_bg")
        st.markdown("**Accents & borders**")
        rate_c = _cp("Rate colour",         "rate_color")
        bord_c = _cp("Border colour",       "border_color")
        bord_w = _fi("Border width (px)",   "border_width", 0, 5, 1)
        cell_p = _fi("Cell padding (px)",   "cell_padding", 2, 20, 6)
        new_style = {"header_bg": h_bg, "header_text": h_text, "header_font": h_font,
                     "header_size": int(h_size), "header_bold": h_bold,
                     "body_font": b_font, "body_size": int(b_size),
                     "body_text": b_text, "body_bg": b_bg, "alt_row_bg": alt_bg,
                     "rate_color": rate_c, "border_color": bord_c,
                     "border_width": int(bord_w), "cell_padding": int(cell_p)}
        if st.button("💾  Save Format", key="save_table_fmt"):
            save_table_style(new_style); st.success("Table format saved.")
        if st.button("Reset to defaults", key="reset_table_fmt"):
            save_table_style(DEFAULT_TABLE_STYLE); st.rerun()
    with prev_col:
        st.markdown("**Live preview**")
        st.markdown(_style_preview_html(new_style), unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("#### Usage Statistics")
    stats = load_stats()
    m1, m2, m3 = st.columns(3)
    m1.metric("Rate Sheets Generated", stats.get("total_rate_sheets", 0))
    m2.metric("Custom Queries Run", stats.get("total_queries", 0))
    m3.metric("Total Actions", stats.get("total_rate_sheets", 0) + stats.get("total_queries", 0))
    events = stats.get("events", [])
    if events:
        st.markdown("**Recent Activity**")
        df_events = pd.DataFrame(list(reversed(events[:50])))
        df_events["Type"] = df_events["Type"].map({
            "rate_sheet": "Rate Sheet Generated",
            "master_query": "Custom Query — Master File",
            "sheet_query": "Custom Query — Formatted Sheet",
        }).fillna(df_events["Type"])
        st.dataframe(df_events, width="stretch", hide_index=True)
    else:
        st.info("No activity recorded yet.")

    st.stop()   # don't render the regular tabs for admin users

# ── Regular tabs (no Admin tab) ────────────────────────────────────────────
tab_data, tab1, tab2, tab3 = st.tabs([
    "Master Data", "Custom Query", "Rate Sheet Generator", "File Format Guide"
])

with tab_data:
    if not _app_settings_store().get("show_master_data", True):
        st.warning("🔒 **Master Data** has been disabled by an administrator.")
        st.stop()
    # ── Save / load shared across all sessions (both CAD and USD) ────────────
    shared = _shared_rate_data()
    save_col, load_col, _ = st.columns([1, 2, 5])
    with save_col:
        if st.button("💾 Save for Team", help="Saves current master rates (CAD & USD) and special rates so any team member can load them."):
            shared["master_grid"]        = st.session_state.master_grid.to_dict(orient="records")
            shared["master_cols"]        = list(st.session_state.master_grid.columns)
            shared["master_grid_usd"]    = st.session_state.master_grid_usd.to_dict(orient="records")
            shared["master_cols_usd"]    = list(st.session_state.master_grid_usd.columns)
            shared["special_rates_v2"]   = st.session_state.special_rates_v2
            shared["special_rates_v2_usd"] = st.session_state.special_rates_v2_usd
            shared["saved_at"]           = datetime.now(_VAN).strftime("%b %d, %Y at %I:%M %p")
            st.success("Saved! Team members can now click 'Use Last' to load this data.")
    with load_col:
        if shared.get("saved_at"):
            if st.button(f"⟳ Use Last  —  {shared['saved_at']}", help="Load the last saved master rates and special rates."):
                st.session_state.master_grid     = pd.DataFrame(
                    shared["master_grid"], columns=shared["master_cols"]
                ).astype(str).fillna("")
                st.session_state.master_grid_usd = pd.DataFrame(
                    shared.get("master_grid_usd", []), columns=shared.get("master_cols_usd", MASTER_GRID_COLS_USD)
                ).astype(str).fillna("") if shared.get("master_grid_usd") else empty_master_df_usd()
                st.session_state.special_rates_v2     = shared.get("special_rates_v2") or []
                st.session_state.special_rates_v2_usd = shared.get("special_rates_v2_usd") or []
                st.rerun()

    st.markdown("---")
    st.subheader("📊 Guaranteed Investment Certificates (CAD)")
    st.caption(
        "Enter your master rates below. "
        "To paste from Excel or Google Sheets: copy your data (Ctrl+C / ⌘C), "
        "click the first cell in the **Issuer** column, then paste (Ctrl+V / ⌘V). "
        "Column order must match the headers shown."
    )
    hdr_col, btn_col = st.columns([8, 1])
    with btn_col:
        if st.button("Clear all", key="clear_cad"):
            st.session_state.master_grid = empty_master_df()
            st.rerun()

    edited = st.data_editor(
        st.session_state.master_grid,
        num_rows="dynamic",
        use_container_width=True,
        height=520,
        column_config={
            "Issuer":                                          st.column_config.TextColumn("Issuer",            width="large"),
            "Credit Rating/Insurance Coverage - Short Term":   st.column_config.TextColumn("CR Short Term",     width="medium"),
            "Credit Rating/Insurance Coverage - Long Term":    st.column_config.TextColumn("CR Long Term",      width="medium"),
            "Take FI Money":                                   st.column_config.TextColumn("Take FI Money",     width="small"),
            "Available":                                       st.column_config.TextColumn("Available",         width="small"),
            "Province":                                        st.column_config.TextColumn("Province",          width="small"),
            "Offers USD?":                                     st.column_config.TextColumn("Offers USD?",       width="small"),
            "As of date for Rates":                            st.column_config.TextColumn("As of Date",        width="medium"),
            "Cashable after 30 days":                          st.column_config.TextColumn("Cash. 30d",         width="medium"),
            "Cashable after 60 days":                          st.column_config.TextColumn("Cash. 60d",         width="medium"),
            "Cashable after 90 days":                          st.column_config.TextColumn("Cash. 90d",         width="medium"),
            "Cashable after 180 days":                         st.column_config.TextColumn("Cash. 180d",        width="medium"),
            "Cashable after 270 days":                         st.column_config.TextColumn("Cash. 270d",        width="medium"),
            "2 Year Cashable after 365 days":                  st.column_config.TextColumn("2Yr Cash. 365d",    width="medium"),
            "30 days":                                         st.column_config.TextColumn("30 days",           width="medium"),
            "60 days":                                         st.column_config.TextColumn("60 days",           width="medium"),
            "90 days":                                         st.column_config.TextColumn("90 days",           width="medium"),
            "120 days":                                        st.column_config.TextColumn("120 days",          width="medium"),
            "180 days":                                        st.column_config.TextColumn("180 days",          width="medium"),
            "270 days":                                        st.column_config.TextColumn("270 days",          width="medium"),
            "1 year fixed":                                    st.column_config.TextColumn("1 Yr Fixed",        width="medium"),
            "18 month fixed":                                  st.column_config.TextColumn("18 Mo Fixed",       width="medium"),
            "2 year fixed":                                    st.column_config.TextColumn("2 Yr Fixed",        width="medium"),
            "3 year fixed":                                    st.column_config.TextColumn("3 Yr Fixed",        width="medium"),
            "4 year fixed":                                    st.column_config.TextColumn("4 Yr Fixed",        width="medium"),
            "5 year fixed":                                    st.column_config.TextColumn("5 Yr Fixed",        width="medium"),
            "Room available":                                  st.column_config.TextColumn("Room Avail.",       width="medium"),
            "Notes":                                           st.column_config.TextColumn("Notes",             width="large"),
        },
    )
    st.session_state.master_grid = edited

    n = master_row_count()
    if n:
        st.caption(f"{n} institution{'s' if n != 1 else ''} entered.")

    st.markdown("---")
    st.subheader("Special Rates")
    st.caption("Enter the institution once, provide its ST and LT credit ratings, then add as many terms as you need.")

    # ── Helper: look up suggested ratings for an issuer ──────────────────────
    def _suggested_ratings(issuer_name):
        store = _rate_history_store()
        norm  = normalize_name(issuer_name)
        li    = (lookup or {}).get(norm)
        def _fmt(r, ins):
            if r and ins: return f"{r} – {ins}"
            return r or ins or ""
        lt = _fmt(li.get("long_rating",""),  li.get("insurance","")) if li else ""
        st_ = _fmt(li.get("short_rating",""), li.get("insurance","")) if li else ""
        return st_, lt

    # ── Add new entry form ───────────────────────────────────────────────────
    with st.form("add_special_rate_form", clear_on_submit=False):
        st.markdown("**Step 1 — Institution**")
        f_issuer = st.text_input("Institution name", key="f_issuer",
                                 placeholder="e.g. EQ Bank")

        st.markdown("**Step 2 — Credit ratings** *(leave blank if none)*")
        fc1, fc2 = st.columns(2)
        with fc1:
            f_st = st.text_input("Short-term rating", key="f_st",
                                 placeholder="e.g. R-1 (High) – CDIC")
        with fc2:
            f_lt = st.text_input("Long-term rating",  key="f_lt",
                                 placeholder="e.g. AA – CDIC")

        st.markdown("**Step 3 — Add a term & rate**")
        ft1, ft2, ft3 = st.columns([3, 2, 1])
        with ft1:
            f_term = st.selectbox("Term", [t[0] for t in TERM_COLUMNS], key="f_term")
        with ft2:
            f_rate = st.text_input("Rate", key="f_rate", placeholder="e.g. 3.75%")
        with ft3:
            st.write("")
            add_term_btn = st.form_submit_button("+ Add Term")

    if add_term_btn:
        issuer_val = st.session_state.get("f_issuer", "").strip()
        rate_val   = st.session_state.get("f_rate",   "").strip()
        term_val   = st.session_state.get("f_term",   "")
        if not issuer_val:
            st.warning("Enter an institution name first.")
        elif not rate_val:
            st.warning("Enter a rate first.")
        else:
            st.session_state.pending_terms.append({"term": term_val, "rate": rate_val})

    # Show pending terms for the current entry
    if st.session_state.pending_terms:
        st.markdown("**Terms being added:**")
        for i, te in enumerate(list(st.session_state.pending_terms)):
            pc1, pc2 = st.columns([5, 1])
            pc1.write(f"• **{te['term']}** — {te['rate']}")
            if pc2.button("✕", key=f"rm_pterm_{i}"):
                st.session_state.pending_terms.pop(i)
                st.rerun()

        issuer_val = st.session_state.get("f_issuer", "").strip()
        st_val     = st.session_state.get("f_st",     "").strip()
        lt_val     = st.session_state.get("f_lt",     "").strip()

        # Auto-suggest ratings if fields are empty
        if issuer_val and not st_val and not lt_val:
            sug_st, sug_lt = _suggested_ratings(issuer_val)
            if sug_st or sug_lt:
                st.caption(f"Suggested — ST: {sug_st or '—'}  |  LT: {sug_lt or '—'}")

        if st.button("✅ Save to Special Rates", key="save_special_entry"):
            if not issuer_val:
                st.warning("Enter an institution name.")
            else:
                entry = {
                    "issuer":    issuer_val,
                    "st_rating": st_val,
                    "lt_rating": lt_val,
                    "entries":   list(st.session_state.pending_terms),
                }
                st.session_state.special_rates_v2.append(entry)
                # Save each term to history
                _term_type_map = {t[0]: t[2] for t in TERM_COLUMNS}
                for te in entry["entries"]:
                    tt     = _term_type_map.get(te["term"], "short")
                    rating = lt_val if tt == "long" else st_val
                    save_rate_to_history(issuer_val, rating, te["term"], te["rate"])
                st.session_state.pending_terms = []
                st.rerun()

    # ── Current special rates list ────────────────────────────────────────────
    if st.session_state.special_rates_v2:
        st.markdown("**Saved special rates:**")
        for i, entry in enumerate(list(st.session_state.special_rates_v2)):
            terms_summary = ", ".join(f"{te['term']} {te['rate']}"
                                      for te in entry.get("entries", []))
            with st.expander(f"**{entry['issuer']}** — {terms_summary}"):
                st.write(f"**ST rating:** {entry.get('st_rating') or '—'}")
                st.write(f"**LT rating:** {entry.get('lt_rating') or '—'}")
                for te in entry.get("entries", []):
                    st.write(f"• {te['term']}: {te['rate']}")
                ec1, ec2, ec3 = st.columns([1, 1, 2])
                with ec1:
                    if st.button("✏️ Edit", key=f"edit_sp_{i}"):
                        st.session_state.edit_special_rate_index = i
                        st.session_state.edit_special_rate_currency = "CAD"
                        st.rerun()
                with ec2:
                    if st.button("🗑 Delete", key=f"del_sp_{i}"):
                        st.session_state.special_rates_v2.pop(i)
                        st.rerun()

        # ── Edit form for CAD special rates ──────────────────────────────────
        if st.session_state.edit_special_rate_index is not None and st.session_state.edit_special_rate_currency == "CAD":
            edit_idx = st.session_state.edit_special_rate_index
            if edit_idx < len(st.session_state.special_rates_v2):
                edit_entry = st.session_state.special_rates_v2[edit_idx]
                st.markdown("---")
                st.markdown(f"**Editing: {edit_entry['issuer']}**")

                # Edit issuer name
                new_issuer = st.text_input("Institution name", value=edit_entry.get('issuer', ''), key="edit_issuer")

                # Edit ratings
                ec1, ec2 = st.columns(2)
                with ec1:
                    new_st = st.text_input("Short-term rating", value=edit_entry.get('st_rating', ''), key="edit_st")
                with ec2:
                    new_lt = st.text_input("Long-term rating", value=edit_entry.get('lt_rating', ''), key="edit_lt")

                # Edit terms
                st.markdown("**Edit terms:**")
                new_entries = []
                for ti, te in enumerate(edit_entry.get('entries', [])):
                    tc1, tc2, tc3 = st.columns([2, 2, 0.5])
                    with tc1:
                        t_term = st.text_input(f"Term {ti+1}", value=te.get('term', ''), key=f"edit_term_{ti}")
                    with tc2:
                        t_rate = st.text_input(f"Rate {ti+1}", value=te.get('rate', ''), key=f"edit_rate_{ti}")
                    with tc3:
                        st.write("")
                        if st.button("✕", key=f"edit_del_term_{ti}"):
                            continue
                    if t_term and t_rate:
                        new_entries.append({"term": t_term, "rate": t_rate})

                # Save changes button
                es1, es2 = st.columns([1, 3])
                with es1:
                    if st.button("✅ Save Changes", key="save_edit_special"):
                        if new_issuer:
                            st.session_state.special_rates_v2[edit_idx] = {
                                "issuer": new_issuer,
                                "st_rating": new_st,
                                "lt_rating": new_lt,
                                "entries": new_entries,
                            }
                            st.session_state.edit_special_rate_index = None
                            st.session_state.edit_special_rate_currency = None
                            st.success("Changes saved!")
                            st.rerun()
                        else:
                            st.warning("Enter an institution name.")
                with es2:
                    if st.button("Cancel", key="cancel_edit_special"):
                        st.session_state.edit_special_rate_index = None
                        st.session_state.edit_special_rate_currency = None
                        st.rerun()

    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.subheader("💵 Guaranteed Investment Certificates (USD)")
    st.markdown("---")
    st.caption(
        "Enter USD GIC rates below. "
        "To paste from Excel or Google Sheets: copy your data (Ctrl+C / ⌘C), "
        "click the first cell in the **Issuer** column, then paste (Ctrl+V / ⌘V). "
        "Column order must match the headers shown."
    )
    usd_hdr_col, usd_btn_col = st.columns([8, 1])
    with usd_btn_col:
        if st.button("Clear all", key="clear_usd"):
            st.session_state.master_grid_usd = empty_master_df_usd()
            st.rerun()

    usd_edited = st.data_editor(
        st.session_state.master_grid_usd,
        num_rows="dynamic",
        use_container_width=True,
        height=520,
        column_config={
            "Issuer":                    st.column_config.TextColumn("Issuer",           width="large"),
            "Available":                 st.column_config.TextColumn("Available",        width="small"),
            "As of date for Rates":      st.column_config.TextColumn("As of Date",      width="medium"),
            "DBRS":                      st.column_config.TextColumn("DBRS",            width="medium"),
            "S&P":                       st.column_config.TextColumn("S&P",             width="medium"),
            "Cashable after 30":         st.column_config.TextColumn("Cash. 30",        width="medium"),
            "Cashable after 90":         st.column_config.TextColumn("Cash. 90",        width="medium"),
            "Cashable after 180":        st.column_config.TextColumn("Cash. 180",       width="medium"),
            "30":                        st.column_config.TextColumn("30",              width="medium"),
            "60":                        st.column_config.TextColumn("60",              width="medium"),
            "90":                        st.column_config.TextColumn("90",              width="medium"),
            "120":                       st.column_config.TextColumn("120",             width="medium"),
            "180":                       st.column_config.TextColumn("180",             width="medium"),
            "270":                       st.column_config.TextColumn("270",             width="medium"),
            "1":                         st.column_config.TextColumn("1 Year",          width="medium"),
            "18 months":                 st.column_config.TextColumn("18 Mo",           width="medium"),
            "2":                         st.column_config.TextColumn("2 Years",         width="medium"),
        },
    )
    st.session_state.master_grid_usd = usd_edited

    n_usd = int(st.session_state.master_grid_usd["Issuer"].astype(str).str.strip().ne("").sum())
    if n_usd:
        st.caption(f"{n_usd} institution{'s' if n_usd != 1 else ''} entered (USD).")

    st.markdown("---")
    st.subheader("Special Rates (USD)")
    st.caption("Enter the institution once, provide its DBRS and S&P ratings, then add as many terms as you need.")

    # ── Add new entry form for USD ────────────────────────────────────────────
    with st.form("add_special_rate_form_usd", clear_on_submit=False):
        st.markdown("**Step 1 — Institution**")
        f_issuer_usd = st.text_input("Institution name", key="f_issuer_usd",
                                      placeholder="e.g. EQ Bank")

        st.markdown("**Step 2 — Credit ratings** *(leave blank if none)*")
        fcu1, fcu2 = st.columns(2)
        with fcu1:
            f_dbrs = st.text_input("DBRS rating", key="f_dbrs",
                                   placeholder="e.g. R-1 (High)")
        with fcu2:
            f_sp = st.text_input("S&P rating", key="f_sp",
                                 placeholder="e.g. A-1+")

        st.markdown("**Step 3 — Add a term & rate**")
        ftu1, ftu2, ftu3 = st.columns([3, 2, 1])
        with ftu1:
            usd_terms = ["Cashable after 30", "Cashable after 90", "Cashable after 180",
                         "30", "60", "90", "120", "180", "270", "1", "18 months", "2"]
            f_term_usd = st.selectbox("Term", usd_terms, key="f_term_usd")
        with ftu2:
            f_rate_usd = st.text_input("Rate", key="f_rate_usd", placeholder="e.g. 3.75%")
        with ftu3:
            st.write("")
            add_term_btn_usd = st.form_submit_button("+ Add Term")

    if add_term_btn_usd:
        issuer_val_usd = st.session_state.get("f_issuer_usd", "").strip()
        rate_val_usd   = st.session_state.get("f_rate_usd",   "").strip()
        term_val_usd   = st.session_state.get("f_term_usd",   "")
        if not issuer_val_usd:
            st.warning("Enter an institution name first.")
        elif not rate_val_usd:
            st.warning("Enter a rate first.")
        else:
            st.session_state.pending_terms.append({"term": term_val_usd, "rate": rate_val_usd})

    # Show pending terms for the current USD entry
    if st.session_state.pending_terms:
        st.markdown("**Terms being added:**")
        for i, te in enumerate(list(st.session_state.pending_terms)):
            pc1, pc2 = st.columns([5, 1])
            pc1.write(f"• **{te['term']}** — {te['rate']}")
            if pc2.button("✕", key=f"rm_pterm_usd_{i}"):
                st.session_state.pending_terms.pop(i)
                st.rerun()

        issuer_val_usd = st.session_state.get("f_issuer_usd", "").strip()
        dbrs_val_usd   = st.session_state.get("f_dbrs",       "").strip()
        sp_val_usd     = st.session_state.get("f_sp",         "").strip()

        if st.button("✅ Save to Special Rates (USD)", key="save_special_entry_usd"):
            if not issuer_val_usd:
                st.warning("Enter an institution name.")
            else:
                entry_usd = {
                    "issuer":    issuer_val_usd,
                    "dbrs":      dbrs_val_usd,
                    "sp":        sp_val_usd,
                    "entries":   list(st.session_state.pending_terms),
                }
                st.session_state.special_rates_v2_usd.append(entry_usd)
                st.session_state.pending_terms = []
                st.rerun()

    # ── Current special rates list (USD) ──────────────────────────────────────
    if st.session_state.special_rates_v2_usd:
        st.markdown("**Saved special rates (USD):**")
        for i, entry in enumerate(list(st.session_state.special_rates_v2_usd)):
            terms_summary = ", ".join(f"{te['term']} {te['rate']}"
                                      for te in entry.get("entries", []))
            with st.expander(f"**{entry['issuer']}** — {terms_summary}"):
                st.write(f"**DBRS:** {entry.get('dbrs') or '—'}")
                st.write(f"**S&P:** {entry.get('sp') or '—'}")
                for te in entry.get("entries", []):
                    st.write(f"• {te['term']}: {te['rate']}")
                eu1, eu2, eu3 = st.columns([1, 1, 2])
                with eu1:
                    if st.button("✏️ Edit", key=f"edit_sp_usd_{i}"):
                        st.session_state.edit_special_rate_index = i
                        st.session_state.edit_special_rate_currency = "USD"
                        st.rerun()
                with eu2:
                    if st.button("🗑 Delete", key=f"del_sp_usd_{i}"):
                        st.session_state.special_rates_v2_usd.pop(i)
                        st.rerun()

        # ── Edit form for USD special rates ──────────────────────────────────
        if st.session_state.edit_special_rate_index is not None and st.session_state.edit_special_rate_currency == "USD":
            edit_idx = st.session_state.edit_special_rate_index
            if edit_idx < len(st.session_state.special_rates_v2_usd):
                edit_entry = st.session_state.special_rates_v2_usd[edit_idx]
                st.markdown("---")
                st.markdown(f"**Editing: {edit_entry['issuer']}**")

                # Edit issuer name
                new_issuer_usd = st.text_input("Institution name", value=edit_entry.get('issuer', ''), key="edit_issuer_usd")

                # Edit ratings
                eu1, eu2 = st.columns(2)
                with eu1:
                    new_dbrs = st.text_input("DBRS rating", value=edit_entry.get('dbrs', ''), key="edit_dbrs")
                with eu2:
                    new_sp = st.text_input("S&P rating", value=edit_entry.get('sp', ''), key="edit_sp")

                # Edit terms
                st.markdown("**Edit terms:**")
                new_entries_usd = []
                for ti, te in enumerate(edit_entry.get('entries', [])):
                    tuc1, tuc2, tuc3 = st.columns([2, 2, 0.5])
                    with tuc1:
                        tu_term = st.text_input(f"Term {ti+1}", value=te.get('term', ''), key=f"edit_term_usd_{ti}")
                    with tuc2:
                        tu_rate = st.text_input(f"Rate {ti+1}", value=te.get('rate', ''), key=f"edit_rate_usd_{ti}")
                    with tuc3:
                        st.write("")
                        if st.button("✕", key=f"edit_del_term_usd_{ti}"):
                            continue
                    if tu_term and tu_rate:
                        new_entries_usd.append({"term": tu_term, "rate": tu_rate})

                # Save changes button
                esu1, esu2 = st.columns([1, 3])
                with esu1:
                    if st.button("✅ Save Changes", key="save_edit_special_usd"):
                        if new_issuer_usd:
                            st.session_state.special_rates_v2_usd[edit_idx] = {
                                "issuer": new_issuer_usd,
                                "dbrs": new_dbrs,
                                "sp": new_sp,
                                "entries": new_entries_usd,
                            }
                            st.session_state.edit_special_rate_index = None
                            st.session_state.edit_special_rate_currency = None
                            st.success("Changes saved!")
                            st.rerun()
                        else:
                            st.warning("Enter an institution name.")
                with esu2:
                    if st.button("Cancel", key="cancel_edit_special_usd"):
                        st.session_state.edit_special_rate_index = None
                        st.session_state.edit_special_rate_currency = None
                        st.rerun()

    # ── Issuer Database ───────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("Issuer Database")
    st.caption("Search for a known issuer to see its ratings and history, then add rates directly.")

    store       = _rate_history_store()
    hist_names  = {v["display_name"] for v in store.values() if "display_name" in v}
    look_names  = {info["display_name"] for info in (lookup or {}).values()}
    all_known   = sorted(hist_names | look_names)

    db_search = st.text_input("Search issuer", placeholder="Start typing…",
                              key="db_search", label_visibility="collapsed")

    if db_search:
        q       = db_search.strip().lower()
        matches = [n for n in all_known if q in n.lower()][:8]
        if not matches:
            st.caption("No known issuers match. Use the form above to add a new one.")
        else:
            for name in matches:
                norm         = normalize_name(name)
                lookup_info  = (lookup or {}).get(norm)
                hist_entries = store.get(name.lower(), {}).get("entries", [])

                def _br(tt):
                    if not lookup_info: return ""
                    r   = lookup_info.get(f"{tt}_rating", "")
                    ins = lookup_info.get("insurance", "")
                    return f"{r} – {ins}" if (r and ins) else (r or ins or "")

                long_r  = _br("long")
                short_r = _br("short")

                with st.expander(f"**{name}**", expanded=(len(matches) == 1)):
                    if lookup_info:
                        if long_r:  st.markdown(f"**LT rating:** {long_r}")
                        if short_r: st.markdown(f"**ST rating:** {short_r}")

                    # Quick-add: term + rate, then saves entry
                    st.markdown("**Add to Special Rates:**")
                    da1, da2, da3 = st.columns([3, 2, 1])
                    with da1:
                        db_term = st.selectbox("Term", [t[0] for t in TERM_COLUMNS],
                                               key=f"db_term_{name}")
                    with da2:
                        db_rate = st.text_input("Rate", key=f"db_rate_{name}",
                                                placeholder="e.g. 3.75%")
                    with da3:
                        st.write("")
                        if st.button("Add", key=f"db_add_{name}"):
                            if db_rate.strip():
                                new_entry = {
                                    "issuer":    name,
                                    "st_rating": short_r,
                                    "lt_rating": long_r,
                                    "entries":   [{"term": db_term, "rate": db_rate.strip()}],
                                }
                                st.session_state.special_rates_v2.append(new_entry)
                                _ttmap = {t[0]: t[2] for t in TERM_COLUMNS}
                                tt     = _ttmap.get(db_term, "short")
                                save_rate_to_history(name, long_r if tt == "long" else short_r,
                                                     db_term, db_rate.strip())
                                st.rerun()
                            else:
                                st.warning("Enter a rate.")

                    if hist_entries:
                        st.markdown("**History:**")
                        for j, he in enumerate(hist_entries[:6]):
                            hc1, hc2, hc3 = st.columns([3, 2, 1])
                            hc1.write(he.get("term", ""))
                            hc2.write(he.get("rate", ""))
                            if hc3.button("Use", key=f"db_use_{name}_{j}"):
                                new_entry = {
                                    "issuer":    name,
                                    "st_rating": short_r,
                                    "lt_rating": long_r,
                                    "entries":   [{"term": he["term"], "rate": he["rate"]}],
                                }
                                st.session_state.special_rates_v2.append(new_entry)
                                st.rerun()


with tab1:
    if not _app_settings_store().get("show_custom_query", True):
        st.warning("🔒 **Custom Query** has been disabled by an administrator.")
        st.stop()

    query_currency = st.radio(
        "Currency",
        ["CAD", "USD"],
        horizontal=True,
        key="query_currency"
    )
    st.write("Filter rates by term, credit rating status, and number of results.")

    query_source = st.radio(
        "Data source",
        ["Master Data", "Formatted Rate Sheet"],
        horizontal=True
    )

    if query_source == "Master Data":
        if query_currency == "CAD":
            n = master_row_count()
            if n:
                st.info(f"{n} institutions loaded from Master Data tab (CAD).")
            else:
                st.warning("No CAD data entered yet — go to the **Master Data** tab and paste your rates.")
        else:  # USD
            n_usd = int(st.session_state.master_grid_usd["Issuer"].astype(str).str.strip().ne("").sum())
            if n_usd:
                st.info(f"{n_usd} institutions loaded from Master Data tab (USD).")
            else:
                st.warning("No USD data entered yet — go to the **Master Data** tab and enter USD rates.")
    else:
        formatted_sheet_file = st.file_uploader(
            "Upload Formatted Rate Sheet",
            type=["xlsx"],
            key="formatted_sheet_uploader"
        )

    # Term options depend on currency
    if query_currency == "CAD":
        term_options = [t[0] for t in TERM_COLUMNS]
    else:  # USD
        usd_terms = ["Cashable after 30", "Cashable after 90", "Cashable after 180",
                     "30", "60", "90", "120", "180", "270", "1", "18 months", "2"]
        term_options = usd_terms

    selected_terms = st.multiselect(
        "Select terms",
        options=term_options,
        default=[]
    )

    col1, col2, col3 = st.columns(3)

    with col1:
        top_n = st.number_input(
            "Top N rates per term",
            min_value=1,
            max_value=50,
            value=3
        )

    with col2:
        min_rate_pct = st.number_input(
            "Minimum rate (%)",
            min_value=0.0,
            max_value=20.0,
            value=0.0,
            step=0.05,
            format="%.2f"
        )

    with col3:
        institution_search = st.text_input(
            "Institution name contains",
            placeholder="e.g. Royal Bank"
        )

    col4, col5, col6 = st.columns(3)

    with col4:
        credit_rated_only = st.checkbox(
            "Credit rated only",
            value=False,
            help="Only show institutions with a formal credit rating (R-1, R-2, AA, BBB, etc.)"
        )

    with col5:
        exclude_cannot_source = st.checkbox(
            "Exclude blank ratings",
            value=False,
            help="Hide rows where no rating or insurance information is available"
        )

    with col6:
        sort_by = st.radio(
            "Sort by",
            options=["rate", "credit"],
            format_func=lambda x: "Highest Rate" if x == "rate" else "Credit Rating",
            horizontal=True
        )

    insurance_filter = st.radio(
        "Insurance filter",
        options=["any", "insured", "cdic", "provincial", "none"],
        format_func=lambda x: {
            "any": "Any", "insured": "Any insured", "cdic": "CDIC only",
            "provincial": "Provincial only", "none": "Uninsured only"
        }[x],
        horizontal=True
    )

    if st.button("Run Query"):
        if not selected_terms:
            st.warning("Please select at least one term.")
        elif query_source == "Master Data" and query_currency == "CAD" and master_row_count() == 0:
            st.error("No CAD data entered — go to the Master Data tab and paste your rates first.")
        elif query_source == "Master Data" and query_currency == "USD" and int(st.session_state.master_grid_usd["Issuer"].astype(str).str.strip().ne("").sum()) == 0:
            st.error("No USD data entered — go to the Master Data tab and enter USD rates first.")
        elif query_source == "Formatted Rate Sheet" and not formatted_sheet_file:
            st.error("Please upload a Formatted Rate Sheet.")
        else:
            # Fetch a large pool (up to 50 per term) so deleted rows can be replaced
            _pool_n = 50
            if query_source == "Formatted Rate Sheet":
                pool = query_from_sheet(formatted_sheet_file, selected_terms, _pool_n, credit_rated_only)
                log_event("sheet_query")
            elif query_currency == "USD":
                # For USD, use master_grid_usd directly
                df_usd = st.session_state.master_grid_usd.copy()
                df_usd = df_usd[df_usd["Issuer"].astype(str).str.strip().ne("")]
                pool = []
                for term in selected_terms:
                    if term not in MASTER_GRID_COLS_USD:
                        continue
                    for _, row in df_usd.iterrows():
                        issuer = row["Issuer"].strip()
                        rating = ""  # USD rates don't have ratings column like CAD
                        rate_str = str(row.get(term, "")).strip()
                        if rate_str:
                            rate = parse_rate(rate_str)
                            if rate >= 0.01:
                                pool.append([issuer, rating, term, rate])
                log_event("master_query")
            else:  # CAD
                pool = generate_custom_query(get_master_file(), lookup, selected_terms, _pool_n, credit_rated_only)
                log_event("master_query")

            # Add special rates (CAD or USD)
            if query_currency == "USD":
                pool = pool + get_special_rate_rows_usd(selected_terms)
            else:
                pool = pool + get_special_rate_rows(selected_terms)
            pool = apply_query_filters(
                pool, min_rate_pct / 100, insurance_filter,
                institution_search.strip(), exclude_cannot_source, sort_by,
            )

            if not pool:
                st.session_state.query_results = None
                st.info("No results matched your query.")
            else:
                st.session_state.query_results = pool
                st.session_state.query_top_n   = int(top_n)
                st.session_state.query_removed  = set()

    if st.session_state.query_results:
        st.markdown("---")

        # ── Commission deduction ──────────────────────────────────────────
        comm_col, _ = st.columns([1, 3])
        with comm_col:
            commission_bps = st.number_input(
                "Commission to deduct (bps)", min_value=0, max_value=200,
                value=0, step=5, help="25 bps turns 4.00% → 3.75%"
            )
        commission = commission_bps / 10000

        # ── Build display: top_n per term, skip removed, pull next-best ───
        from collections import defaultdict as _dd
        pool    = st.session_state.query_results
        removed = st.session_state.query_removed
        top_n_d = st.session_state.get("query_top_n", 3)

        # Preserve original term order
        seen_terms = []
        for r in pool:
            if r[2] not in seen_terms:
                seen_terms.append(r[2])

        pool_by_term = _dd(list)
        for r in pool:
            pool_by_term[r[2]].append(r)

        display_rows = []
        for term in seen_terms:
            count = 0
            for r in pool_by_term[term]:
                if count >= top_n_d:
                    break
                if (r[0], r[2]) in removed:
                    continue
                display_rows.append(r)
                count += 1

        # Apply commission
        display_rows_adj = [
            (r[0], r[1], r[2], max(0.0, r[3] - commission))
            for r in display_rows
        ]

        st.markdown("**Results** — click ✕ to remove a row; the next best replaces it:")
        for i, row in enumerate(display_rows_adj):
            issuer, rating, term, rate = row
            c1, c2, c3, c4, c5 = st.columns([3, 4, 2, 1, 0.4])
            c1.write(issuer)
            c2.write(rating)
            c3.write(term)
            c4.write(f"{rate * 100:.2f}%")
            if c5.button("✕", key=f"del_row_{i}_{issuer}_{term}"):
                st.session_state.query_removed.add((issuer, term))
                st.rerun()

        if display_rows_adj:
            _copy_button_component(
                build_copy_html(display_rows_adj),
                "Copy to Clipboard",
            )

with tab2:
    if not _app_settings_store().get("show_rate_sheet", True):
        st.warning("🔒 **Rate Sheet Generator** has been disabled by an administrator.")
        st.stop()

    rs_currency = st.radio(
        "Currency",
        ["CAD", "USD"],
        horizontal=True,
        key="rs_currency"
    )

    if rs_currency == "CAD":
        n         = master_row_count()
        n_special = len(get_special_rate_rows())
    else:  # USD
        n         = int(st.session_state.master_grid_usd["Issuer"].astype(str).str.strip().ne("").sum())
        n_special = len(get_special_rate_rows_usd())

    has_data  = n > 0 or n_special > 0

    if n:
        st.info(
            f"{n} institutions in Master Data ({rs_currency})."
            + (f"  ·  {n_special} special rate{'s' if n_special != 1 else ''}." if n_special else "")
        )
    elif n_special:
        st.info(f"{n_special} special rate{'s' if n_special != 1 else ''} entered ({rs_currency}).")
    else:
        st.warning(f"No {rs_currency} data yet — go to the **Master Data** tab and paste your rates.")

    def _build_and_store(key, fi_only=False, credit_only=False):
        if not has_data:
            st.error("No data entered — add master rates or special rates in the Master Data tab first.")
            return
        if rs_currency == "CAD":
            base    = generate_report(get_master_file(), lookup, fi_only=fi_only) if n > 0 else []
            special = get_special_rate_rows()
        else:  # USD
            # For USD, build from master_grid_usd
            df_usd = st.session_state.master_grid_usd.copy()
            df_usd = df_usd[df_usd["Issuer"].astype(str).str.strip().ne("")]
            base = []
            for _, row in df_usd.iterrows():
                issuer = row["Issuer"].strip()
                for term in MASTER_GRID_COLS_USD[1:]:  # Skip Issuer column
                    if term in ["Available", "As of date for Rates", "DBRS", "S&P"]:
                        continue
                    rate_str = str(row.get(term, "")).strip()
                    if rate_str:
                        rate = parse_rate(rate_str)
                        if rate >= 0.01:
                            # USD rows have no rating (DBRS/S&P are separate)
                            base.append([issuer, "", term, rate])
            special = get_special_rate_rows_usd()

        if credit_only:
            base    = [r for r in base    if is_credit_or_guarantee(r[1])]
            special = [r for r in special if is_credit_or_guarantee(r[1])]
        output = sort_output(base + special)

        # Filter by visible terms (only for "All In GIC Rates")
        if key == "rs_all_in_html":
            output = [r for r in output if r[2] in st.session_state.visible_terms]

        if fi_only:
            # FI tables show only the credit rating; 100%/insurance-only → em dash
            output = [[r[0], fi_rating(r[1]), r[2], r[3]] for r in output]
        st.session_state[key] = build_copy_html(output)
        log_event("rate_sheet")

    # ── 1. All In GIC Rates ───────────────────────────────────────────────────
    st.subheader("All In GIC Rates")
    st.caption("Every available rate from the master data, sorted by term then rate.")

    # Term visibility toggles
    with st.expander("⚙️ Show/hide terms", expanded=False):
        all_terms = [t[0] for t in TERM_COLUMNS]
        visible = st.session_state.visible_terms

        st.markdown("**Fixed Terms**")
        fixed_terms = [t for t in all_terms if "Cashable" not in t]
        ft_col1, ft_col2, ft_col3 = st.columns(3)
        cols = [ft_col1, ft_col2, ft_col3]
        for idx, term in enumerate(fixed_terms):
            col = cols[idx % 3]
            with col:
                if st.checkbox(term, value=term in visible, key=f"vis_{term}"):
                    visible.add(term)
                else:
                    visible.discard(term)

        st.markdown("**Cashable Terms**")
        cash_terms = [t for t in all_terms if "Cashable" in t]
        ct_col1, ct_col2, ct_col3 = st.columns(3)
        cols = [ct_col1, ct_col2, ct_col3]
        for idx, term in enumerate(cash_terms):
            col = cols[idx % 3]
            with col:
                if st.checkbox(term, value=term in visible, key=f"vis_{term}"):
                    visible.add(term)
                else:
                    visible.discard(term)

        if st.button("Reset to defaults", key="reset_visible_terms"):
            st.session_state.visible_terms = {
                "5 Year Fixed", "4 Year Fixed", "3 Year Fixed", "2 Year Fixed",
                "18 Month Fixed", "1 Year Fixed", "270 Day Fixed", "180 Day Fixed",
                "90 Day Fixed", "60 Day Fixed", "30 Day Fixed",
                "1 Year Cashable After 90 Days", "1 Year Cashable After 30 Days"
            }
            st.rerun()

    if st.button("Generate — All In GIC Rates"):
        _build_and_store("rs_all_in_html")
    if st.session_state.get("rs_all_in_html"):
        _copy_button_component(st.session_state.rs_all_in_html, "Copy — All In GIC Rates")

    st.markdown("---")

    # ── 2. Credit Rated & 100% Guarantees (CAD only) ────────────────────────
    if rs_currency == "CAD":
        st.subheader("Credit Rated & 100% Guarantees Only")
        st.caption(
            "Only institutions with a formal credit rating (R-1, R-2, AA, BBB …) "
            "or a 100% guarantee. All unrated and uninsured rows are removed."
        )
        if st.button("Generate — Credit Rated & 100% Guarantees"):
            _build_and_store("rs_credit_html", credit_only=True)
        if st.session_state.get("rs_credit_html"):
            _copy_button_component(st.session_state.rs_credit_html, "Copy — Credit Rated & Guarantees")

    st.markdown("---")

    # ── Full Morning Email ────────────────────────────────────────────────────
    st.subheader("Full Morning Email")
    st.caption("Auto-generates email with GIC CAD (live) + GIC USD (live) + HISA placeholders.")

    if st.button("Generate Full Email"):
        try:
            has_cad_data = master_row_count() > 0 or len(get_special_rate_rows()) > 0
            if not has_cad_data:
                st.error("No CAD GIC data entered — add master rates in the Master Data tab first.")
            else:
                # Build GIC CAD data (live)
                base_cad   = generate_report(get_master_file(), lookup) if master_row_count() > 0 else []
                gic_cad    = sort_output(base_cad + get_special_rate_rows())

                # Build GIC USD data (live)
                n_usd = int(st.session_state.master_grid_usd["Issuer"].astype(str).str.strip().ne("").sum())
                gic_usd = None
                if n_usd > 0:
                    df_usd = st.session_state.master_grid_usd.copy()
                    df_usd = df_usd[df_usd["Issuer"].astype(str).str.strip().ne("")]
                    gic_usd_base = []
                    for _, row in df_usd.iterrows():
                        issuer = row["Issuer"].strip()
                        for term in MASTER_GRID_COLS_USD[1:]:
                            if term in ["Available", "As of date for Rates", "DBRS", "S&P"]:
                                continue
                            rate_str = str(row.get(term, "")).strip()
                            if rate_str:
                                rate = parse_rate(rate_str)
                                if rate >= 0.01:
                                    gic_usd_base.append([issuer, "", term, rate])
                    gic_usd_special = get_special_rate_rows_usd()
                    gic_usd = sort_output(gic_usd_base + gic_usd_special) if gic_usd_base or gic_usd_special else None

                # Get template and styles from admin settings and substitute
                settings = _app_settings_store()
                template = settings.get("email_template", _DEFAULT_EMAIL_TEMPLATE)
                email_font = settings.get("email_font", "Calibri")
                email_font_size = settings.get("email_font_size", 11)
                email_text_color = settings.get("email_text_color", "#000000")
                table_style = load_table_style()

                st.session_state.full_email_html = build_email_from_template(
                    template, gic_cad, gic_usd,
                    style=table_style,
                    email_font=email_font,
                    email_font_size=email_font_size,
                    email_text_color=email_text_color
                )
                log_event("rate_sheet")
        except Exception as e:
            st.error(f"Error generating email: {str(e)}")

    if st.session_state.get("full_email_html"):
        _copy_button_component(st.session_state.full_email_html, "Copy Full Email")


with tab3:
    st.subheader("Master Rates File — Required Format")
    st.write(
        "The file must be **.xlsx**. Column names are case-insensitive. "
        "One row per institution. Columns marked **Used** are read by the app; others are accepted but ignored."
    )

    master_cols = pd.DataFrame([
        {"Column Name": "Issuer",                                  "Used by App": "Yes", "Description": "Name of the institution (must be in column A).", "Example": "Royal Bank of Canada"},
        {"Column Name": "Insurance/Credit Rating Short Term",      "Used by App": "No",  "Description": "For reference only — ratings are managed in the backend lookup.", "Example": "R-1 (High)"},
        {"Column Name": "Insurance/Credit Rating Long Term",       "Used by App": "No",  "Description": "For reference only — ratings are managed in the backend lookup.", "Example": "AA"},
        {"Column Name": "Take FI Money",                           "Used by App": "Yes", "Description": "Set to 'Yes' to include this institution in the FI Rate Table.", "Example": "Yes"},
        {"Column Name": "Available",                               "Used by App": "Yes", "Description": "Must say 'available' for the row to be included. Any other value skips it.", "Example": "available"},
        {"Column Name": "Province",                                "Used by App": "No",  "Description": "Internal reference field.", "Example": "ON"},
        {"Column Name": "Offers USD",                              "Used by App": "No",  "Description": "Internal reference field.", "Example": "Yes"},
        {"Column Name": "As of Date",                              "Used by App": "No",  "Description": "Date the rates were last updated.", "Example": "2024-01-15"},
        {"Column Name": "Cashable After 30 Days",                  "Used by App": "Yes", "Description": "Rate for 1 Year Cashable After 30 Days term.", "Example": "3.05%"},
        {"Column Name": "Cashable After 90 Days",                  "Used by App": "Yes", "Description": "Rate for 1 Year Cashable After 90 Days term.", "Example": "3.10%"},
        {"Column Name": "30 Days",                                 "Used by App": "Yes", "Description": "Rate for 30 Day Fixed term.", "Example": "2.80%"},
        {"Column Name": "60 Days",                                 "Used by App": "Yes", "Description": "Rate for 60 Day Fixed term.", "Example": "2.90%"},
        {"Column Name": "90 Days",                                 "Used by App": "Yes", "Description": "Rate for 90 Day Fixed term.", "Example": "3.00%"},
        {"Column Name": "180 Days",                                "Used by App": "Yes", "Description": "Rate for 180 Day Fixed term.", "Example": "3.10%"},
        {"Column Name": "270 Days",                                "Used by App": "Yes", "Description": "Rate for 270 Day Fixed term.", "Example": "3.15%"},
        {"Column Name": "1 Year Fixed",                            "Used by App": "Yes", "Description": "Rate for 1 Year Fixed term.", "Example": "3.20%"},
        {"Column Name": "18 Month Fixed",                          "Used by App": "Yes", "Description": "Rate for 18 Month Fixed term.", "Example": "3.30%"},
        {"Column Name": "2 Year Fixed",                            "Used by App": "Yes", "Description": "Rate for 2 Year Fixed term.", "Example": "3.40%"},
        {"Column Name": "3 Year Fixed",                            "Used by App": "Yes", "Description": "Rate for 3 Year Fixed term.", "Example": "3.50%"},
        {"Column Name": "4 Year Fixed",                            "Used by App": "Yes", "Description": "Rate for 4 Year Fixed term.", "Example": "3.60%"},
        {"Column Name": "5 Year Fixed",                            "Used by App": "Yes", "Description": "Rate for 5 Year Fixed term.", "Example": "3.75%"},
    ])

    st.dataframe(master_cols.set_index("Column Name").T, width="stretch")
    st.caption("Rates can be entered as percentages (3.75%) or decimals (0.0375). Blank cells or rates below 1% are ignored.")


st.markdown("---")
st.caption(
    "This tool does not guarantee 100% accuracy. Rates and information should be independently verified before use. "
    "For any issues, please contact Henry Jagger at henryjagger@gmail.com."
)
